import streamlit as st
import pandas as pd
import io
import json
import os
import openpyxl
from geopy.distance import geodesic
import openpyxl, os, datetime, numpy as np, io
import time
import base64

# ============================
# 🔵 SPLASH SCREEN CON IMAGEN
# ============================
if "splash_done" not in st.session_state:
    st.session_state["splash_done"] = False

if not st.session_state["splash_done"]:

    image_path = "splash.png"

    with open(image_path, "rb") as img_file:
        encoded_string = base64.b64encode(img_file.read()).decode()

    splash_css = f"""
    <style>
    .splash-container {{
        position: fixed;
        top:0; left:0;
        width:100%; height:100%;
        background-color:#0A1128;
        display:flex;
        justify-content:center;
        align-items:center;
        flex-direction:column;
        z-index:999999;
    }}
    .splash-logo {{
        width: 950px;           
        max-width: 95%;          
        border-radius: 8px;
        box-shadow: 0 0 25px rgba(0,0,0,0.45);
    }}
    .loader {{
        margin-top: 28px;
        border: 6px solid #f3f3f3;
        border-top: 6px solid #00AEEF;
        border-radius: 50%;
        width: 58px;
        height: 58px;
        animation: spin 1s linear infinite;
    }}
    @keyframes spin {{
        0% {{ transform: rotate(0deg); }}
        100% {{ transform: rotate(360deg); }}
    }}
    .splash-text {{
        color:white;
        margin-top:20px;
        font-size:22px;
        font-family:'Segoe UI';
    }}
    </style>
    """

    splash_html = f"""
    <div class="splash-container">
        <img src="data:image/jpeg;base64,{encoded_string}" class="splash-logo">
        <div class="loader"></div>
        <div class="splash-text">Cargando SMIpim...</div>
    </div>
    """

    st.markdown(splash_css, unsafe_allow_html=True)
    st.markdown(splash_html, unsafe_allow_html=True)

    time.sleep(2.3)
    st.session_state["splash_done"] = True
    st.rerun()

# ============================
# 🔧 CONFIGURACIÓN DE RUTAS
# ============================

import sys, os

# Detectar si la app se ejecuta desde un .exe o desde el navegador
if getattr(sys, 'frozen', False):
    # Cuando corre como ejecutable, usar la carpeta donde está el .exe
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Cuando corre desde Python/Streamlit, usar la carpeta del script
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Función para acceder a los archivos dentro de la app
def ruta_archivo(nombre):
    """Devuelve la ruta absoluta de un archivo dentro de la carpeta base."""
    return os.path.join(BASE_DIR, nombre)


# ---------------- CONFIGURACIÓN ----------------
st.set_page_config(page_title="Selección de Mercados", layout="wide")

# Ruta donde se guardarán los usuarios
ruta_usuarios = "usuarios.json"

# ---------------- CARGA DE USUARIOS ----------------
# Si no existe el archivo, se crea con los valores iniciales
if os.path.exists(ruta_usuarios):
    with open(ruta_usuarios, "r", encoding="utf-8") as f:
        usuarios_validos = json.load(f)
else:
    usuarios_validos = {
        "admin": {"password": "1234", "rol": "admin"},
        "usuario": {"password": "1234", "rol": "usuario"}
    }
    with open(ruta_usuarios, "w", encoding="utf-8") as f:
        json.dump(usuarios_validos, f, ensure_ascii=False, indent=4)

# ---------------- SESIÓN ----------------
if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False
if "usuario" not in st.session_state:
    st.session_state["usuario"] = ""
if "rol" not in st.session_state:
    st.session_state["rol"] = ""
if "tabla_editada" not in st.session_state:
    st.session_state["tabla_editada"] = None
if "archivo_excel_bytes" not in st.session_state:
    st.session_state["archivo_excel_bytes"] = None
if "productos_filtrados_tab1" not in st.session_state:
    st.session_state["productos_filtrados_tab1"] = pd.DataFrame()

# ---------------- FUNCIONES AUXILIARES ----------------
def obtener_coordenadas_desde_excel(pais, df_paises):
    """Obtiene coordenadas desde la hoja Paises"""
    fila = df_paises[df_paises["Nombre"].str.lower().str.strip() == pais.lower().strip()]
    if not fila.empty:
        lat, lon = fila.iloc[0]["Latitud"], fila.iloc[0]["Longitud"]
        if pd.notna(lat) and pd.notna(lon):
            return (lat, lon)
    return None

def calcular_distancia(pais1, pais2, df_paises):
    """Calcula distancia geográfica entre dos países"""
    coord1 = obtener_coordenadas_desde_excel(pais1, df_paises)
    coord2 = obtener_coordenadas_desde_excel(pais2, df_paises)
    if not coord1 or not coord2:
        return None
    try:
        return geodesic(coord1, coord2).kilometers
    except Exception:
        return None

def calcular_costo_transporte(pais_destino, pais_origen, df_paises, tarifa_km=0.38):
    """Costo de transporte entre países (contenedor 40ft, mercancía general)"""
    coord_origen = obtener_coordenadas_desde_excel(pais_origen, df_paises)
    coord_destino = obtener_coordenadas_desde_excel(pais_destino, df_paises)
    if not coord_origen or not coord_destino:
        return None
    try:
        distancia = geodesic(coord_origen, coord_destino).kilometers
        if distancia > 0:
            return round(distancia * tarifa_km, 2)
    except Exception as e:
        print(f"Error al calcular distancia entre {pais_origen} y {pais_destino}: {e}")
    return None

def guardar_en_excel_sin_tocar_otras_hojas(ruta_original, hoja_objetivo, df, ruta_salida):
    """
    Guarda los datos de `df` en la hoja indicada, manteniendo intactas las demás hojas del Excel original.
    Crea un nuevo archivo con las modificaciones (ruta_salida).
    """
    try:
        # Cargar el libro original
        libro = openpyxl.load_workbook(ruta_original)

        # Si la hoja existe, se borra su contenido sin eliminar las demás
        if hoja_objetivo in libro.sheetnames:
            hoja = libro[hoja_objetivo]
            hoja.delete_rows(1, hoja.max_row)
        else:
            hoja = libro.create_sheet(hoja_objetivo)

        # Escribir encabezados
        for col_idx, col_name in enumerate(df.columns, start=1):
            hoja.cell(row=1, column=col_idx, value=col_name)

        # Escribir datos
        for row_idx, fila in enumerate(df.itertuples(index=False), start=2):
            for col_idx, valor in enumerate(fila, start=1):
                hoja.cell(row=row_idx, column=col_idx, value=valor)

        # Guardar como nuevo archivo
        libro.save(ruta_salida)
        return True
    except Exception as e:
        st.error(f"Error al guardar en Excel: {e}")
        return False

# ---------------- LOGIN ----------------
if not st.session_state["autenticado"]:
    st.title("Inicio de sesión")

    # Campos de inicio de sesión normales
    usuario = st.text_input("Usuario")
    contraseña = st.text_input("Contraseña", type="password")

    if st.button("Iniciar sesión"):
        if usuario in usuarios_validos and usuarios_validos[usuario]["password"] == contraseña:
            st.session_state["autenticado"] = True
            st.session_state["usuario"] = usuario
            st.session_state["rol"] = usuarios_validos[usuario]["rol"]
            st.success(f"Bienvenido, {usuario} (Rol: {st.session_state['rol'].capitalize()})")
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos.")

    # ---------------- BOTÓN CREAR NUEVO USUARIO ----------------
    st.markdown("---")
    st.subheader("¿Quieres registrar un nuevo usuario?")
    if "mostrar_login_admin" not in st.session_state:
        st.session_state["mostrar_login_admin"] = False
    if "admin_autenticado" not in st.session_state:
        st.session_state["admin_autenticado"] = False

    # Botón que activa la verificación del admin
    if st.button("Crear nuevo usuario"):
        st.session_state["mostrar_login_admin"] = True

    # Si el usuario presionó el botón, pedirle login del admin
    if st.session_state["mostrar_login_admin"] and not st.session_state["admin_autenticado"]:
        st.info("Debes iniciar sesión como administrador para continuar.")
        admin_user = st.text_input("Usuario administrador")
        admin_pass = st.text_input("Contraseña administrador", type="password")

        if st.button("Verificar administrador"):
            if (
                admin_user in usuarios_validos
                and usuarios_validos[admin_user]["password"] == admin_pass
                and usuarios_validos[admin_user]["rol"] == "admin"
            ):
                st.success("Administrador verificado correctamente.")
                st.session_state["admin_autenticado"] = True
            else:
                st.error("Usuario o contraseña incorrectos. Solo el administrador puede continuar.")

    # Si ya fue verificado el admin, mostrar formulario para crear nuevo usuario
    if st.session_state["admin_autenticado"]:
        with st.expander("Formulario de registro de nuevo usuario", expanded=True):
            nuevo_usuario = st.text_input("Nuevo usuario")
            nueva_contraseña = st.text_input("Nueva contraseña", type="password")
            nuevo_rol = st.selectbox("Rol del usuario", ["usuario", "admin"])

            if st.button("Registrar usuario"):
                if nuevo_usuario and nueva_contraseña:
                    if nuevo_usuario not in usuarios_validos:
                        usuarios_validos[nuevo_usuario] = {
                            "password": nueva_contraseña,
                            "rol": nuevo_rol
                        }
                        with open(ruta_usuarios, "w", encoding="utf-8") as f:
                            json.dump(usuarios_validos, f, ensure_ascii=False, indent=4)
                        st.success(f"Usuario '{nuevo_usuario}' creado correctamente con rol '{nuevo_rol}'.")
                        st.session_state["mostrar_login_admin"] = False
                        st.session_state["admin_autenticado"] = False
                    else:
                        st.warning("Ese usuario ya existe.")
                else:
                    st.error("Debes ingresar ambos campos.")
else:
    # ---------------- SESIÓN ACTIVA ----------------
    st.sidebar.success(f"Usuario: {st.session_state['usuario']} ({st.session_state['rol']})")
    if st.sidebar.button("Cerrar sesión"):
        st.session_state["autenticado"] = False
        st.session_state["usuario"] = ""
        st.session_state["rol"] = ""
        st.rerun()

    st.title("Aplicativo Selección de Mercados Internacionales")

    # ---------------- PESTAÑAS ----------------
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs([
        "Productos", "Costo (COST)", "Logística (LOGI)",
        "Comercial (COMM)", "Economía (ECON)", "Política (POLI)",
        "Cultura (CULT)", "Visualización de Tablas Totales", "Gráficos"
    ])

    # ------------------ PESTAÑA 1 ------------------
    with tab1:
        st.title("Productos por categoría y subcategoría")
        # ============================================================
        # SELECTOR UNIVERSAL DEL PAÍS DESTINO (para COST, LOGI y TAB8)
        # ============================================================
        st.markdown("Selecciona el país destino de importación")

        # Cargar países desde el Excel principal
        ruta_paises = "2025-09-25T04-20_export_con_todos_los_paises_actualizado.xlsx"
        try:
            df_paises_dest = pd.read_excel(ruta_paises, sheet_name="Paises")
            lista_destinos = sorted(df_paises_dest["Nombre"].astype(str).unique())
        except:
            lista_destinos = ["Colombia"]  # fallback si no existe hoja

        # Selector
        pais_destino = st.selectbox(
            "País destino",
            lista_destinos,
            index=lista_destinos.index("Colombia") if "Colombia" in lista_destinos else 0,
            key="selector_destino"
        )

        # Guardar en session_state
        st.session_state["pais_destino"] = pais_destino

        st.success(f"País destino seleccionado: {pais_destino}")

        archivo_excel = st.file_uploader("Carga el archivo Excel", type=["xlsx"], key="excel_tab1")

        if archivo_excel:
            # Guardar archivo subido en memoria
            st.session_state["archivo_excel_bytes"] = archivo_excel.read()
            archivo_excel.seek(0)

            import os, unicodedata, io, re

            # ------------------ Carga del archivo ------------------

            ruta_modificada = "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx"

            if not os.path.exists(ruta_modificada):
                st.error("No existe el archivo modificado. Primero debes guardarlo desde la pestaña.")
                st.stop()

            # Cargar siempre desde el modificado
            productos = pd.read_excel(ruta_modificada, sheet_name="Codigos productos y subproducto")
            precios = pd.read_excel(ruta_modificada, sheet_name="Precios")


            # ------------------ Funciones auxiliares ------------------
            def normalize(s):
                s = str(s)
                s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
                return s.lower().strip().replace("-", " ").replace("  ", " ")


            def format_codigo(x):
                if pd.isna(x):
                    return ""
                try:
                    return f"{int(float(str(x).strip())):02d}"
                except:
                    return str(x).zfill(2)

            # ------------------ Normalización ------------------
            productos["CodigoProductoFmt"] = productos["CodigoProducto"].apply(format_codigo)
            productos["CodigoSubproductoFmt"] = productos["CodigoSubproducto"].apply(format_codigo)

            productos["NombreProducto_norm"] = productos["NombreProducto"].astype(str).map(normalize)
            productos["NombreSubproducto_norm"] = productos["NombreSubproducto"].astype(str).map(normalize)
            precios["producto_norm"] = precios["producto"].astype(str).map(normalize)

            productos["Categoria"] = productos["CodigoProductoFmt"] + " - " + productos["NombreProducto"]
            productos["Subcategoria"] = productos["CodigoSubproductoFmt"] + " - " + productos["NombreSubproducto"]

            categorias = ["Todos"] + sorted(
                productos.drop_duplicates(subset=["CodigoProductoFmt"])["Categoria"].tolist()
            )

            # ------------------ Filtros laterales ------------------
            st.sidebar.markdown("### Filtros de búsqueda")
            categoria_sel = st.sidebar.selectbox("Selecciona una categoría", categorias, key="cat_sel_tab1")

            if categoria_sel != "Todos":
                codigo_categoria, _ = categoria_sel.split(" - ", 1)
                sub_df = productos[productos["CodigoProductoFmt"] == codigo_categoria]
                subcategorias = ["Todos"] + sorted(sub_df["Subcategoria"].dropna().unique())
            else:
                subcategorias = ["Todos"] + sorted(productos["Subcategoria"].dropna().unique())

            subcat_sel = st.sidebar.selectbox("Selecciona una subcategoría", subcategorias, key="subcat_sel_tab1")

            # ------------------ Campos búsqueda ------------------
            st.sidebar.markdown("### Búsquedas personalizadas")
            nombre_input = st.sidebar.text_input("Buscar por nombre", key="nombre_tab1").strip()
            codigo_input = st.sidebar.text_input("Buscar por código (2 dígitos)", key="codigo_tab1").strip()
            subcodigo_input = st.sidebar.text_input("Buscar por subcódigo (2 dígitos)", key="subcodigo_tab1").strip()

            # ------------------ FILTROS APLICADOS ------------------
            productos_filtrados = precios.copy()

            # ---------- FILTRO POR NOMBRE ----------
            if nombre_input:
                nombre_norm = normalize(nombre_input)
                productos_filtrados = productos_filtrados[
                    productos_filtrados["producto_norm"].str.contains(nombre_norm, na=False)
                ]

            # ---------- FILTRO POR CÓDIGO ----------
            if codigo_input:
                codigo_input = codigo_input.zfill(2)
                nombres_cat = productos.loc[
                    productos["CodigoProductoFmt"] == codigo_input,
                    "NombreProducto_norm"
                ].unique()

                tokens = set()
                for n in nombres_cat:
                    for w in n.split():
                        if len(w) > 2:
                            tokens.add(w)

                productos_filtrados = productos_filtrados[
                    productos_filtrados["producto_norm"].apply(lambda x: any(t in x for t in tokens))
                ]

            # ---------- FILTRO POR SUBCÓDIGO ----------
            if subcodigo_input:
                subcodigo_input = subcodigo_input.zfill(2)
                nombres_sub = productos.loc[
                    productos["CodigoSubproductoFmt"] == subcodigo_input,
                    "NombreSubproducto_norm"
                ].unique()

                tokens = set()
                for n in nombres_sub:
                    palabras = re.split(r"[ \-,/()]+", n)
                    for w in palabras:
                        w = normalize(w)
                        if len(w) > 2:
                            tokens.add(w)

                tokens_expand = set(tokens)
                for t in tokens:
                    if t.endswith("s"):
                        tokens_expand.add(t[:-1])
                    else:
                        tokens_expand.add(t + "s")

                productos_filtrados = productos_filtrados[
                    productos_filtrados["producto_norm"].apply(lambda x: any(t in x for t in tokens_expand))
                ]

            #FILTRO DE CATEGORÍA)

            if categoria_sel != "Todos":
                codigo_cat = categoria_sel.split(" - ")[0]

                # Todas las subcategorías que pertenecen a esta categoría
                mask_subs = productos["CodigoProductoFmt"] == codigo_cat
                subcodigos = productos.loc[mask_subs, "CodigoSubproductoFmt"].dropna().unique()

                tokens_cat = set()

                # Generar tokens desde los subproductos reales
                for sc in subcodigos:
                    nombres_sub = productos.loc[
                        productos["CodigoSubproductoFmt"] == sc,
                        "NombreSubproducto_norm"
                    ].dropna().unique()

                    for n in nombres_sub:
                        palabras = re.split(r"[ \-,/()]+", n)
                        for w in palabras:
                            w = normalize(w)
                            if len(w) > 2:
                                tokens_cat.add(w)

                # Expandir singular/plural
                tokens_expand_cat = set(tokens_cat)
                for t in tokens_cat:
                    if t.endswith("s"):
                        tokens_expand_cat.add(t[:-1])
                    else:
                        tokens_expand_cat.add(t + "s")

                # Filtrar productos reales
                productos_filtrados = productos_filtrados[
                    productos_filtrados["producto_norm"].apply(lambda x: any(t in x for t in tokens_expand_cat))
                ]

            # FILTRO DE SUBCATEGORÍA

            if subcat_sel != "Todos":
                codigo_sub = subcat_sel.split(" - ")[0]

                nombres_sub = productos.loc[
                    productos["CodigoSubproductoFmt"] == codigo_sub,
                    "NombreSubproducto_norm"
                ].dropna().unique()

                tokens = set()
                for n in nombres_sub:
                    palabras = re.split(r"[ \-,/()]+", n)
                    for w in palabras:
                        w = normalize(w)
                        if len(w) > 2:
                            tokens.add(w)

                tokens_expand = set(tokens)
                for t in tokens:
                    if t.endswith("s"):
                        tokens_expand.add(t[:-1])
                    else:
                        tokens_expand.add(t + "s")

                productos_filtrados = productos_filtrados[
                    productos_filtrados["producto_norm"].apply(lambda x: any(t in x for t in tokens_expand))
                ]

            # ------------------ RESULTADOS ------------------
            if not productos_filtrados.empty:
                productos_filtrados["precio"] = pd.to_numeric(productos_filtrados["precio"], errors="coerce")
                productos_ordenados = productos_filtrados.sort_values(by="precio", na_position="last")

                if categoria_sel != "Todos":
                    st.markdown(f"**Categoría seleccionada:** {categoria_sel}")
                if subcat_sel != "Todos":
                    st.markdown(f"**Subcategoría seleccionada:** {subcat_sel}")

                st.subheader("Resultados de búsqueda")

                # ------------------ CRUD ------------------
                st.markdown("### Gestión de Datos (Productos filtrados)")
                col1, col2, col3 = st.columns(3)
                ruta_original = "2025-09-25T04-20_export_con_todos_los_paises_actualizado.xlsx"
                ruta_salida = "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx"

                # ---- AÑADIR ----
                with col1:
                    with st.expander("Añadir producto"):
                        pais = st.text_input("País", key="pais_add_tab1")
                        producto = st.text_input("Producto", key="producto_add_tab1")
                        precio = st.number_input("Precio", min_value=0.0, key="precio_add_tab1")

                        if st.button("Guardar nuevo producto", key="btn_add_tab1"):
                            if pais and producto:
                                nuevo = pd.DataFrame([[pais, producto, precio]],
                                                     columns=["pais", "producto", "precio"])

                                df_actual = pd.read_excel(ruta_original, sheet_name="Precios")
                                df_actual = pd.concat([df_actual, nuevo], ignore_index=True)

                                guardar_en_excel_sin_tocar_otras_hojas(
                                    ruta_original,
                                    "Precios",
                                    df_actual,
                                    ruta_salida
                                )

                                # RECARGAR DESDE EL ARCHIVO MODIFICADO
                                precios = pd.read_excel(ruta_salida, sheet_name="Precios")

                                st.session_state["productos_filtrados_tab1"] = precios.copy()
                                st.success("Producto añadido correctamente.")
                                st.rerun()
                            else:
                                st.warning("Debes ingresar país y producto.")

                # ---- EDITAR ----
                with col2:
                    with st.expander("Editar producto existente"):
                        df = productos_ordenados.copy()
                        if not df.empty:
                            opciones = {f"{r['pais']} – {r['producto']} (${r['precio']})": i for i, r in df.iterrows()}
                            seleccion = st.selectbox("Selecciona producto a editar", list(opciones.keys()),
                                                     key="fila_edit_tab1")

                            fila = opciones[seleccion]
                            nuevo_pais = st.text_input("Nuevo país", value=df.loc[fila, "pais"], key="pais_edit_tab1")
                            nuevo_producto = st.text_input("Nuevo producto", value=df.loc[fila, "producto"],
                                                           key="producto_edit_tab1")
                            nuevo_precio = st.number_input("Nuevo precio", value=float(df.loc[fila, "precio"]),
                                                           min_value=0.0, key="precio_edit_tab1")

                            if st.button("Actualizar producto", key="btn_edit_tab1"):
                                df_actual = pd.read_excel(ruta_original, sheet_name="Precios")

                                df_actual.at[fila, "pais"] = nuevo_pais
                                df_actual.at[fila, "producto"] = nuevo_producto
                                df_actual.at[fila, "precio"] = nuevo_precio

                                guardar_en_excel_sin_tocar_otras_hojas(
                                    ruta_original,
                                    "Precios",
                                    df_actual,
                                    ruta_salida
                                )

                                # Recargar desde el archivo modificado
                                precios = pd.read_excel(ruta_salida, sheet_name="Precios")

                                st.session_state["productos_filtrados_tab1"] = precios.copy()
                                st.success("Producto actualizado correctamente.")
                                st.rerun()

                # ---- ELIMINAR ----
                with col3:
                    with st.expander("Eliminar producto existente"):
                        df = productos_ordenados.copy()

                        if not df.empty:

                            # Crear opciones estilo "pais – producto ($precio)"
                            opciones = {
                                f"{r['pais']} – {r['producto']} (${r['precio']})": i
                                for i, r in df.iterrows()
                            }

                            # Selector
                            seleccion = st.selectbox(
                                "Selecciona producto a eliminar",
                                list(opciones.keys()),
                                key="fila_delete_tab1"
                            )

                            # Obtener índice de la tabla filtrada
                            fila = opciones[seleccion]

                            # Obtener valores reales de esa fila
                            pais_val = df.loc[fila, "pais"]
                            producto_val = df.loc[fila, "producto"]
                            precio_val = df.loc[fila, "precio"]

                            # Botón para eliminar
                            if st.button("Eliminar producto", key="btn_delete_tab1_exec"):
                                df_actual = pd.read_excel(ruta_original, sheet_name="Precios")

                                # Eliminar por VALORES, no por índice
                                df_actual = df_actual[
                                    ~(
                                            (df_actual["pais"] == pais_val) &
                                            (df_actual["producto"] == producto_val) &
                                            (df_actual["precio"] == precio_val)
                                    )
                                ].reset_index(drop=True)

                                guardar_en_excel_sin_tocar_otras_hojas(
                                    ruta_original,
                                    "Precios",
                                    df_actual,
                                    ruta_salida
                                )

                                # Recargar modificado
                                precios = pd.read_excel(ruta_salida, sheet_name="Precios")
                                st.session_state["productos_filtrados_tab1"] = precios.copy()

                                st.warning("Producto eliminado correctamente.")
                                st.rerun()

                        else:
                            st.info("No hay productos para eliminar.")

                # ------------------ DESCARGA ------------------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    productos_ordenados.to_excel(writer, index=False, sheet_name="Productos")

                st.download_button(
                    label="Descargar Excel actualizado",
                    data=output.getvalue(),
                    file_name="Productos_actualizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_tab1"
                )

                # Mostrar tabla final
                st.dataframe(productos_ordenados[["pais", "producto", "precio"]].reset_index(drop=True))

                # Guardar para Tab2
                st.session_state["productos_filtrados_tab1"] = productos_ordenados.copy()

                # Forzar recálculo de COSTO en Tab2
                st.session_state["df_cost"] = None
                st.session_state["force_reload_tab2"] = True

            else:
                st.info("No hay productos disponibles para estos filtros.")

        else:
            st.info("Por favor, carga un archivo Excel para comenzar.")

        # ------------------ PESTAÑA 2 ------------------
        with tab2:
            st.title("1.Costo (COST) — Estandarización de criterios")

            # Forzar recálculo si Tab1 cambió los productos
            if st.session_state.get("force_reload_tab2"):
                st.session_state["force_reload_tab2"] = False
                st.session_state["df_cost"] = None

            # ---------------------------
            # LECTURA ORIGINAL DEL EXCEL (NO SE MODIFICA, LO USA EL CRUD)
            # ---------------------------
            productos_filtrados = pd.read_excel(
                "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx",
                sheet_name="Precios"
            )

            # 🔥 LÍNEA AÑADIDA — ahora TAB2 usa el filtro de TAB1 si existe
            productos_filtrados = st.session_state.get("productos_filtrados_tab1", productos_filtrados)

            if not productos_filtrados.empty:

                import os, openpyxl, datetime

                ruta_excel = "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx"

                if not os.path.exists(ruta_excel):
                    st.error(
                        "❌ No se encontró el archivo Excel modificado. Guarda primero los datos desde la pestaña 1.")
                    st.stop()


                # FUNCIÓN UNIVERSAL NUMÉRICA
                def convertir_numero_seguro(valor):
                    if valor is None:
                        return None

                    valor = str(valor).strip()

                    if valor.lower() in ["none", "nan", "", "n/a", "null"]:
                        return None

                    simbolos = ["$", "usd", "₡", "crc", "%"]
                    for s in simbolos:
                        valor = valor.replace(s, "")

                    valor = valor.replace(" ", "")

                    if valor.count(",") == 1 and valor.count(".") > 1:
                        partes = valor.split(",")
                        izquierda = partes[0].replace(".", "")
                        valor = izquierda + "." + partes[1]
                    elif valor.count(".") == 1 and valor.count(",") > 1:
                        valor = valor.replace(",", "")
                    elif valor.count(",") == 1 and valor.count(".") == 0:
                        valor = valor.replace(",", ".")

                    try:
                        return float(valor)
                    except:
                        return None


                # ============================================================
                # 1. PPAO DESDE EL PRODUCTO FILTRADO
                # ============================================================
                resumen = (productos_filtrados.groupby("pais")
                           .agg({"precio": "mean"})
                           .rename(columns={"precio": "Precio del producto en origen (PPAO)"})
                           .reset_index())
                resumen.rename(columns={"pais": "Paises"}, inplace=True)

                resumen["PPAO_float"] = resumen["Precio del producto en origen (PPAO)"].apply(convertir_numero_seguro)

                resumen = resumen.sort_values(
                    by="PPAO_float",
                    ascending=True,
                    na_position="last"
                ).head(300).reset_index(drop=True)

                # ============================================================
                # 2. INTC AUTOMÁTICO
                # ============================================================
                paises_excel = pd.read_excel(ruta_excel, sheet_name="Paises")
                paises_excel["Nombre"] = paises_excel["Nombre"].astype(str).str.strip()

                pais_base = st.session_state.get("pais_destino", None)

                if not pais_base:
                    st.error("Debes seleccionar un país destino en la pestaña 1.")
                    st.stop()

                st.success(f"País destino seleccionado para cálculos de COST: **{pais_base}**")


                def intc_seguro(dest):
                    if dest.strip().lower() == pais_base.strip().lower():
                        return None
                    return calcular_costo_transporte(dest, pais_base, paises_excel)


                resumen["Costos de transporte internacional (INTC)"] = \
                    resumen["Paises"].apply(intc_seguro)

                # ============================================================
                # 3. CEBC — COSTO FRONTERIZO
                # ============================================================
                precios_export = pd.read_excel(ruta_excel, sheet_name="PreciosExportacion")
                precios_export.columns = precios_export.columns.str.strip()

                if "País" in precios_export.columns:
                    precios_export.rename(columns={"País": "Paises"}, inplace=True)
                elif "Pais" in precios_export.columns:
                    precios_export.rename(columns={"Pais": "Paises"}, inplace=True)

                precios_export["Paises"] = precios_export["Paises"].astype(str).str.strip()

                col_cebc = None
                for col in precios_export.columns:
                    if "costo" in col.lower() and "usd" in col.lower():
                        col_cebc = col
                        break

                if col_cebc is None:
                    st.error("❌ No existe una columna CEBC válida en PreciosExportacion.")
                    st.stop()

                precios_export[col_cebc] = precios_export[col_cebc].apply(convertir_numero_seguro)
                precios_export = precios_export.dropna(subset=[col_cebc])

                map_cebc = precios_export.groupby("Paises", as_index=False)[col_cebc].mean()

                resumen["Costo de exportación del cumplimiento fronterizo (CEBC)"] = \
                    resumen["Paises"].map(map_cebc.set_index("Paises")[col_cebc])

                resumen["Costo de exportación del cumplimiento fronterizo (CEBC)"] = \
                    resumen["Costo de exportación del cumplimiento fronterizo (CEBC)"].apply(
                        lambda x: None if x in [0, 0.0] else x
                    )

                # ============================================================
                # CRUD (SIN CAMBIAR NADA)
                # ============================================================
                st.markdown("### 🔧 Gestión de Datos (Hoja PreciosExportacion)")
                col1, col2 = st.columns(2)

                # EDITAR
                with col1:
                    with st.expander("Editar costo existente"):

                        df_edit = pd.read_excel(ruta_excel, sheet_name="Precios")
                        df_edit.columns = df_edit.columns.str.strip().str.lower()

                        if "pais" not in df_edit.columns:
                            st.error("La hoja 'Precios' no tiene columna 'pais'.")
                            st.stop()

                        if not df_edit.empty:

                            paises = sorted(df_edit["pais"].astype(str).unique())

                            pais_sel = st.selectbox("Selecciona país", paises, key="edit_pais_cost")

                            df_pe = pd.read_excel(ruta_excel, sheet_name="PreciosExportacion")
                            df_pe.columns = df_pe.columns.str.strip().str.lower()

                            col_pais_pe = None
                            for c in df_pe.columns:
                                if c in ["paises", "pais", "país"]:
                                    col_pais_pe = c
                                    break

                            col_cebc_pe = None
                            for c in df_pe.columns:
                                if "costo" in c or "cebc" in c:
                                    col_cebc_pe = c
                                    break

                            if col_pais_pe is None or col_cebc_pe is None:
                                st.error("La hoja PreciosExportacion no tiene columnas válidas.")
                                st.stop()

                            match = df_pe[df_pe[col_pais_pe].str.lower().str.strip() == pais_sel.lower().strip()]
                            costo_actual = float(match[col_cebc_pe].values[0]) if not match.empty else 0.0

                            nuevo_costo = st.number_input(
                                "Nuevo costo (USD)",
                                value=costo_actual,
                                min_value=0.0,
                                step=0.1,
                                key="edit_costo_cost"
                            )

                            if st.button("Actualizar costo (PreciosExportacion)", key="btn_edit_cost"):

                                wb = openpyxl.load_workbook(ruta_excel)
                                hoja = wb["PreciosExportacion"]

                                idx_pais = None
                                idx_cebc = None
                                for idx, celda in enumerate(hoja[1]):
                                    nombre = str(celda.value).strip().lower() if celda.value else ""
                                    if nombre in ["paises", "pais", "país"]:
                                        idx_pais = idx
                                    if "costo" in nombre or "cebc" in nombre:
                                        idx_cebc = idx

                                encontrado = False

                                for fila in hoja.iter_rows(min_row=2, values_only=False):
                                    pais_celda = str(fila[idx_pais].value).strip().lower() if fila[
                                        idx_pais].value else ""
                                    if pais_celda == pais_sel.strip().lower():
                                        fila[idx_cebc].value = nuevo_costo
                                        encontrado = True
                                        break

                                if not encontrado:
                                    hoja.append([pais_sel, nuevo_costo])

                                try:
                                    wb.save(ruta_excel)
                                except PermissionError:
                                    base, ext = os.path.splitext(ruta_excel)
                                    timestamp = datetime.datetime.now().strftime("%H%M%S")
                                    ruta_copia = f"{base}_copia_{timestamp}{ext}"
                                    wb.save(ruta_copia)
                                    st.warning(f"Archivo abierto. Se guardó como copia: {ruta_copia}")

                                st.success(f"Costo actualizado correctamente.")
                                st.rerun()

                        else:
                            st.info("No hay datos para editar.")

                # ELIMINAR
                with col2:
                    with st.expander("Eliminar país"):

                        df_del = pd.read_excel(ruta_excel, sheet_name="Precios")
                        df_del.columns = df_del.columns.str.strip().str.lower()

                        if "pais" not in df_del.columns:
                            st.error("La hoja 'Precios' no tiene columna 'pais'.")
                            st.stop()

                        if not df_del.empty:

                            paises = sorted(df_del["pais"].astype(str).unique())

                            pais_sel = st.selectbox("Selecciona país", paises, key="delete_pais_cost")

                            if st.button("Eliminar país (PreciosExportacion)", key="btn_del_cost"):

                                wb = openpyxl.load_workbook(ruta_excel)
                                hoja = wb["PreciosExportacion"]

                                filas_a_eliminar = []
                                for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=False), start=2):
                                    if str(fila[0].value).strip().lower() == pais_sel.strip().lower():
                                        filas_a_eliminar.append(i)

                                for i in reversed(filas_a_eliminar):
                                    hoja.delete_rows(i, 1)

                                try:
                                    wb.save(ruta_excel)
                                except PermissionError:
                                    base, ext = os.path.splitext(ruta_excel)
                                    timestamp = datetime.datetime.now().strftime("%H%M%S")
                                    ruta_copia = f"{base}_copia_{timestamp}{ext}"
                                    wb.save(ruta_copia)
                                    st.warning(f"Archivo abierto. Se guardó como copia: {ruta_copia}")

                                st.warning(f"País '{pais_sel}' eliminado correctamente.")
                                st.rerun()

                        else:
                            st.info("No hay países para eliminar.")

                # ============================================================
                # TABLA BASE
                # ============================================================
                st.subheader("Tabla de costos base")

                df_base = resumen.copy()
                df_base["PPAO_float"] = df_base["PPAO_float"].apply(convertir_numero_seguro)
                df_base["INTC_float"] = df_base["Costos de transporte internacional (INTC)"].apply(
                    convertir_numero_seguro)
                df_base["CEBC_float"] = df_base["Costo de exportación del cumplimiento fronterizo (CEBC)"].apply(
                    convertir_numero_seguro)

                for col in ["PPAO_float", "INTC_float", "CEBC_float"]:
                    df_base[col] = df_base[col].apply(lambda x: None if x in [0, 0.0] else x)

                df_base["__faltantes"] = df_base[["PPAO_float", "INTC_float", "CEBC_float"]].isna().sum(axis=1)

                df_base = df_base.sort_values(
                    by=["__faltantes", "PPAO_float", "INTC_float", "CEBC_float"],
                    ascending=[True, True, True, True]
                ).drop(columns=["__faltantes"])

                st.dataframe(
                    df_base[[
                        "Paises",
                        "Precio del producto en origen (PPAO)",
                        "Costos de transporte internacional (INTC)",
                        "Costo de exportación del cumplimiento fronterizo (CEBC)"
                    ]],
                    use_container_width=True
                )

                # ============================================================
                # TABLA NORMALIZADA
                # ============================================================
                st.markdown("---")
                st.subheader("Costos normalizados")

                df = resumen.copy()
                A3 = 10

                df["PDPO_num"] = df_base["PPAO_float"]
                df["CDTI_num"] = df_base["INTC_float"]
                df["CEBC_num"] = df_base["CEBC_float"]

                min_pdpo = df["PDPO_num"].min()
                min_cdti = df["CDTI_num"].min()


                def norm(valor, minimo):
                    if valor is None or minimo is None:
                        return None
                    if valor <= 0 or minimo <= 0:
                        return None
                    return round(A3 * (minimo / valor), 2)


                df["PDPO_norm"] = df["PDPO_num"].apply(lambda x: norm(x, min_pdpo))
                df["CDTI_norm"] = df["CDTI_num"].apply(lambda x: norm(x, min_cdti))

                cebc_validos = df["CEBC_num"].dropna()
                cebc_validos = cebc_validos[cebc_validos > 0]

                if not cebc_validos.empty:
                    min_cebc = cebc_validos.min()
                    df["CEBC_norm"] = df["CEBC_num"].apply(
                        lambda x: round(A3 * (min_cebc / x), 2) if x not in [None, 0] else None
                    )
                else:
                    df["CEBC_norm"] = None

                PESO_PDPO = 0.35
                PESO_CDTI = 0.35
                PESO_CEBC = 0.30


                def costo_total(row):
                    p1 = row["PDPO_norm"] if row["PDPO_norm"] is not None else 0
                    p2 = row["CDTI_norm"] if row["CDTI_norm"] is not None else 0
                    p3 = row["CEBC_norm"] if row["CEBC_norm"] is not None else 0
                    return round((PESO_PDPO * p1) + (PESO_CDTI * p2) + (PESO_CEBC * p3), 2)


                df["Costo_Total_Normalizado"] = df.apply(costo_total, axis=1)

                df["__faltantes"] = df[[
                    "PDPO_norm", "CDTI_norm", "CEBC_norm", "Costo_Total_Normalizado"
                ]].isna().sum(axis=1)

                df = df.sort_values(
                    by=["__faltantes", "Costo_Total_Normalizado", "PDPO_norm", "CDTI_norm", "CEBC_norm"],
                    ascending=[True, True, True, True, True]
                ).drop(columns=["__faltantes"])

                st.dataframe(
                    df[[
                        "Paises",
                        "PDPO_norm",
                        "CDTI_norm",
                        "CEBC_norm",
                        "Costo_Total_Normalizado"
                    ]],
                    use_container_width=True
                )

                # Guardar resultado final
                st.session_state["df_cost"] = df.copy()

            else:
                st.info("ℹ No hay productos filtrados en la pestaña 'Productos'.")

    # ------------------ PESTAÑA 3 -------------------------
    with tab3:
        st.title("2. Logística (LOGI)")

        # ================= BOTONES CRUD =================
        st.markdown("### 🔧 Gestión de Datos (Tabla LOGI)")
        col1, col2, col3 = st.columns(3)

        import os, io
        from geopy.distance import geodesic

        ruta_excel = "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx"

        if not os.path.exists(ruta_excel):
            st.warning("⚠️ No se encontró el archivo Excel con los datos logísticos.")
        else:
            try:
                xls = pd.ExcelFile(ruta_excel, engine="openpyxl")

                # Detectar la hoja LOGI
                hojas = {h.lower(): h for h in xls.sheet_names}
                hoja_logi = next((v for k, v in hojas.items() if "logi" in k), None)

                if not hoja_logi:
                    st.warning("No existe hoja LOGI. Se creará vacía temporalmente.")
                    df_logi = pd.DataFrame(columns=[
                        "Paises",
                        "Índice de desempeño logístico (LPIN)",
                        "Tráfico del puerto de contenedores (CPT)",
                        "Tiempo de tránsito del transporte internacional (ITTT)"
                    ])
                else:
                    df_logi = pd.read_excel(xls, sheet_name=hoja_logi)

                    # Normalizar nombres
                    rename_map = {}
                    for c in df_logi.columns:
                        cl = str(c).strip().lower()

                        if cl in ["pais", "paises"]:
                            rename_map[c] = "Paises"
                        elif "lpi" in cl:
                            rename_map[c] = "Índice de desempeño logístico (LPIN)"
                        elif "container" in cl or "cpt" in cl:
                            rename_map[c] = "Tráfico del puerto de contenedores (CPT)"
                        elif "transit" in cl or "ittt" in cl:
                            rename_map[c] = "Tiempo de tránsito del transporte internacional (ITTT)"

                    df_logi = df_logi.rename(columns=rename_map)

                    # Garantizar columnas
                    for col_req in [
                        "Paises",
                        "Índice de desempeño logístico (LPIN)",
                        "Tráfico del puerto de contenedores (CPT)",
                        "Tiempo de tránsito del transporte internacional (ITTT)"
                    ]:
                        if col_req not in df_logi.columns:
                            df_logi[col_req] = None

                    df_logi["Paises"] = df_logi["Paises"].astype(str).str.strip()

                st.session_state["tabla_logi"] = df_logi.copy()

                # ----------------------- CRUD AÑADIR -----------------------
                with col1:
                    with st.expander("Añadir país"):
                        new_pais = st.text_input("País nuevo", key="logi_new_pais")
                        new_lpin = st.number_input("LPIN", min_value=0.0, step=0.01, key="logi_new_lpin")
                        new_cpt = st.number_input("CPT", min_value=0.0, step=1.0, key="logi_new_cpt")

                        if st.button("Guardar país LOGI", key="btn_add_logi"):
                            wb = openpyxl.load_workbook(ruta_excel)
                            if hoja_logi not in wb.sheetnames:
                                hoja = wb.create_sheet("Logistica (LOGI)")
                            else:
                                hoja = wb[hoja_logi]

                            hoja.append([new_pais, new_lpin, new_cpt, None])  # ITTT lo calcula la app

                            try:
                                wb.save(ruta_excel)
                            except PermissionError:
                                base, ext = os.path.splitext(ruta_excel)
                                ts = datetime.datetime.now().strftime("%H%M%S")
                                ruta_copia = f"{base}_copia_{ts}{ext}"
                                wb.save(ruta_copia)
                                st.warning(f"Archivo abierto. Guardado como copia: {ruta_copia}")

                            st.success(f"País '{new_pais}' añadido correctamente.")
                            st.rerun()

                # ----------------------- CRUD EDITAR -----------------------
                with col2:
                    with st.expander("Editar país"):
                        df_edit = st.session_state["tabla_logi"].copy()

                        if not df_edit.empty:
                            pais_sel = st.selectbox("Selecciona país", df_edit["Paises"].tolist(), key="logi_edit_pais")

                            fila_sel = df_edit[df_edit["Paises"] == pais_sel].iloc[0]

                            new_lpin = st.number_input(
                                "Nuevo LPIN",
                                value=float(fila_sel["Índice de desempeño logístico (LPIN)"]) if fila_sel[
                                    "Índice de desempeño logístico (LPIN)"] else 0.0,
                                min_value=0.0, step=0.1
                            )

                            new_cpt = st.number_input(
                                "Nuevo CPT",
                                value=float(fila_sel["Tráfico del puerto de contenedores (CPT)"]) if fila_sel[
                                    "Tráfico del puerto de contenedores (CPT)"] else 0.0,
                                min_value=0.0, step=1.0
                            )

                            if st.button("Actualizar LOGI", key="btn_edit_logi"):
                                wb = openpyxl.load_workbook(ruta_excel)
                                hoja = wb[hoja_logi]

                                idx_pais = idx_lpin = idx_cpt = None
                                for i, celda in enumerate(hoja[1]):
                                    nombre = str(celda.value).strip().lower()
                                    if nombre in ["paises", "pais"]:
                                        idx_pais = i
                                    if "lpi" in nombre:
                                        idx_lpin = i
                                    if "container" in nombre or "cpt" in nombre:
                                        idx_cpt = i

                                for fila in hoja.iter_rows(min_row=2, values_only=False):
                                    if str(fila[idx_pais].value).strip().lower() == pais_sel.lower():
                                        fila[idx_lpin].value = new_lpin
                                        fila[idx_cpt].value = new_cpt
                                        break

                                try:
                                    wb.save(ruta_excel)
                                except PermissionError:
                                    base, ext = os.path.splitext(ruta_excel)
                                    ts = datetime.datetime.now().strftime("%H%M%S")
                                    ruta_copia = f"{base}_copia_{ts}{ext}"
                                    wb.save(ruta_copia)
                                    st.warning(f"Archivo abierto. Guardado como copia: {ruta_copia}")

                                st.success("País actualizado correctamente.")
                                st.rerun()

                # ----------------------- CRUD ELIMINAR -----------------------
                with col3:
                    with st.expander("Eliminar país"):
                        df_del = st.session_state["tabla_logi"].copy()

                        if not df_del.empty:
                            pais_sel = st.selectbox("Selecciona país", df_del["Paises"].tolist(), key="logi_del_pais")

                            if st.button("Eliminar de LOGI", key="btn_del_logi"):
                                wb = openpyxl.load_workbook(ruta_excel)
                                hoja = wb[hoja_logi]

                                filas_eliminar = []
                                for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=False), start=2):
                                    if str(fila[0].value).strip().lower() == pais_sel.lower():
                                        filas_eliminar.append(i)

                                for i in reversed(filas_eliminar):
                                    hoja.delete_rows(i, 1)

                                try:
                                    wb.save(ruta_excel)
                                except PermissionError:
                                    base, ext = os.path.splitext(ruta_excel)
                                    ts = datetime.datetime.now().strftime("%H%M%S")
                                    ruta_copia = f"{base}_copia_{ts}{ext}"
                                    wb.save(ruta_copia)
                                    st.warning(f"Archivo abierto. Guardado como copia: {ruta_copia}")

                                st.warning(f"País '{pais_sel}' eliminado correctamente.")
                                st.rerun()

                # ======================================================
                # CARGAR PUERTOS
                # ======================================================
                df_puertos = pd.read_excel(
                    io.BytesIO(st.session_state["archivo_excel_bytes"]),
                    sheet_name="Puertos"
                )
                df_puertos["País"] = df_puertos["País"].astype(str).str.strip()
                df_puertos["Puerto"] = df_puertos["Puerto"].astype(str).str.strip()

                # ======================================================
                # PAÍS DESTINO AUTOMÁTICO DESDE TAB 1
                # ======================================================
                pais_origen = st.session_state.get("pais_destino", None)

                if not pais_origen:
                    st.error("⚠️ Debes seleccionar un país destino en TAB 1 para calcular logística.")
                    st.stop()

                st.success(f"País destino seleccionado automáticamente: **{pais_origen}**")

                df_temp = st.session_state["tabla_logi"].copy()

                # Puerto base del país destino
                puerto_base = df_puertos[df_puertos["País"] == pais_origen].iloc[0]
                coord_origen = (puerto_base["Latitud"], puerto_base["Longitud"])

                # velocidad estandar barco
                velocidad_kmh = 18 * 1.852

                # Calcular ITTT para cada país productor
                for i, row in df_temp.iterrows():
                    pais_dest = str(row["Paises"]).strip()
                    puertos_dest = df_puertos[df_puertos["País"] == pais_dest]
                    if puertos_dest.empty:
                        continue

                    puerto_dest = puertos_dest.iloc[0]
                    coord_dest = (puerto_dest["Latitud"], puerto_dest["Longitud"])

                    distancia_km = geodesic(coord_origen, coord_dest).kilometers
                    tiempo_dias = round(distancia_km / (velocidad_kmh * 24), 2)

                    df_temp.at[i, "Tiempo de tránsito del transporte internacional (ITTT)"] = f"{tiempo_dias} días"

                st.session_state["tabla_logi"] = df_temp.copy()
                st.success("ITTT calculado automáticamente según el país destino seleccionado.")

                #  TABLA BASE
                df_base = st.session_state["tabla_logi"].copy()


                def to_float_ittt(x):
                    try:
                        return float(str(x).replace(" días", "").replace(",", "."))
                    except:
                        return None


                df_base["__completos"] = df_base[
                    [
                        "Índice de desempeño logístico (LPIN)",
                        "Tráfico del puerto de contenedores (CPT)",
                        "Tiempo de tránsito del transporte internacional (ITTT)"
                    ]
                ].notna().all(axis=1)

                df_base["ITTT_float"] = df_base["Tiempo de tránsito del transporte internacional (ITTT)"].apply(
                    to_float_ittt
                )

                df_logi_ordenada = df_base.sort_values(
                    by=["__completos",
                        "Índice de desempeño logístico (LPIN)",
                        "Tráfico del puerto de contenedores (CPT)",
                        "ITTT_float"],
                    ascending=[False, False, False, True],
                    na_position="last"
                ).drop(columns=["__completos", "ITTT_float"])

                st.subheader("Tabla LOGI (Datos Base)")
                st.dataframe(df_logi_ordenada, use_container_width=True)

            except Exception as e:
                st.error(f"Error al cargar datos logísticos: {e}")

        # TABLA NORMALIZADA (MISMA FÓRMULA QUE TU EXCEL)
        st.markdown("---")
        st.subheader("Tabla Logística Normalizada (LOGI)")

        df_rank = st.session_state.get("tabla_logi", pd.DataFrame()).copy()

        if not df_rank.empty:

            def to_float(x):
                try:
                    v = float(str(x).replace(",", ".").replace(" días", "").strip())
                    return v if v > 0 else None
                except:
                    return None


            df_rank["LPIN"] = df_rank["Índice de desempeño logístico (LPIN)"].apply(to_float)
            df_rank["CPT"] = df_rank["Tráfico del puerto de contenedores (CPT)"].apply(to_float)
            df_rank["ITTT"] = df_rank["Tiempo de tránsito del transporte internacional (ITTT)"].apply(to_float)

            A3 = 10

            MAX_LPIN_CONST = 4.3
            MAX_CPT_CONST = 278_982_714.2857
            MIN_ITTT_CONST = 0.58744

            df_rank["LPIN_norm"] = df_rank["LPIN"].apply(
                lambda x: round(A3 * x / MAX_LPIN_CONST, 2) if x and MAX_LPIN_CONST else None
            )
            df_rank["CPT_norm"] = df_rank["CPT"].apply(
                lambda x: round(A3 * x / MAX_CPT_CONST, 2) if x and MAX_CPT_CONST else None
            )
            df_rank["ITTT_norm"] = df_rank["ITTT"].apply(
                lambda x: round(A3 * MIN_ITTT_CONST / x, 2) if x and MIN_ITTT_CONST else None
            )


            def calcular_total(row):
                if (
                        row["LPIN_norm"] is None or
                        row["CPT_norm"] is None or
                        row["ITTT_norm"] is None
                ):
                    return None
                return round(
                    0.30 * row["LPIN_norm"] +
                    0.30 * row["CPT_norm"] +
                    0.40 * row["ITTT_norm"],
                    2
                )


            df_rank["Puntaje_LOGI"] = df_rank.apply(calcular_total, axis=1)

            df_rank["__completos"] = df_rank[
                ["LPIN_norm", "CPT_norm", "ITTT_norm"]
            ].notna().all(axis=1)

            df_rank = df_rank.sort_values(
                by=["__completos", "Puntaje_LOGI"],
                ascending=[False, False]
            ).drop(columns=["__completos"]).reset_index(drop=True)

            columnas_finales = ["Paises", "LPIN_norm", "CPT_norm", "ITTT_norm", "Puntaje_LOGI"]

            st.dataframe(
                df_rank[columnas_finales].rename(columns={"Puntaje_LOGI": "Costo_Total_Logistico_Normalizado"}),
                use_container_width=True
            )

            st.session_state["df_logi"] = df_rank.rename(
                columns={"Puntaje_LOGI": "Costo_Total_Logistico_Normalizado"}
            ).copy()

        else:
            st.info("No hay datos logísticos cargados todavía.")

    # ------------------ PESTAÑA 4 ------------------
    with tab4:
        st.title("3. Comercial (COMM)")

        import openpyxl, os, datetime, numpy as np

        # Inicializar overlay
        if "comm_overrides" not in st.session_state:
            st.session_state["comm_overrides"] = pd.DataFrame(columns=[
                "Paises",
                "Índice de penetración en el mercado de exportación (IEMP)",
                "Índice de Libertad Económica (IOEF)"
            ])

        def append_valor_excel(ruta_excel: str, hoja: str, pais: str, valor):
            if not os.path.exists(ruta_excel):
                st.error(f"No se encontró el archivo: {ruta_excel}")
                return False

            wb = openpyxl.load_workbook(ruta_excel)
            if hoja not in wb.sheetnames:
                wb.create_sheet(hoja)
            ws = wb[hoja]

            if ws.max_row == 1 and all(ws.cell(row=1, column=c).value is None for c in range(1, 3)):
                ws.cell(row=1, column=1, value="País")
                ws.cell(row=1, column=2, value="Valor")

            ws.append([pais, valor])

            try:
                wb.save(ruta_excel)
                wb.close()
                return True
            except PermissionError:
                base, ext = os.path.splitext(ruta_excel)
                copia = f"{base}_copia_{datetime.datetime.now().strftime('%H%M%S')}{ext}"
                wb.save(copia)
                wb.close()
                st.warning(f"El archivo estaba abierto. Se guardó una copia: {copia}")
                return True


        st.markdown("### Gestión de Datos (COMM)")
        col1, col2, col3 = st.columns(3)

        # ===================== AÑADIR =====================
        with col1:
            with st.expander("Añadir país"):
                pais_add = st.text_input("País", key="pais_add_tab4")
                iemp_add = st.number_input("Índice de penetración (IEMP)", min_value=0.0, key="iemp_add_tab4")
                ioef_add = st.number_input("Índice de libertad económica (IOEF)", min_value=0.0, key="ioef_add_tab4")

                if st.button("Guardar país (COMM)", key="btn_add_tab4"):
                    ruta_excel = os.path.abspath("2025-09-25T04-20_export_con_todos_los_paises_actualizado.xlsx")

                    ovr = st.session_state["comm_overrides"].copy()
                    registro = {
                        "Paises": pais_add,
                        "Índice de penetración en el mercado de exportación (IEMP)": iemp_add,
                        "Índice de Libertad Económica (IOEF)": ioef_add
                    }
                    ovr = pd.concat([ovr, pd.DataFrame([registro])], ignore_index=True)
                    st.session_state["comm_overrides"] = ovr

                    append_valor_excel(ruta_excel, "Índice de penetración mercado", pais_add, iemp_add)
                    append_valor_excel(ruta_excel, "Indice de Libertad Económica", pais_add, ioef_add)

                    st.success(f"País '{pais_add}' añadido correctamente al Excel.")
                    st.rerun()

        # ===================== EDITAR =====================
        with col2:
            with st.expander("Editar país"):
                df_edit_src = st.session_state.get("comm_overrides", pd.DataFrame())

                if not df_edit_src.empty:
                    opciones = {r["Paises"]: i for i, r in df_edit_src.iterrows()}
                    pais_sel = st.selectbox("Selecciona país a editar", list(opciones.keys()),
                                            key="pais_edit_select_tab4")
                    fila_edit = opciones[pais_sel]


                    def _to_float(v):
                        try:
                            return float(v)
                        except:
                            return 0.0


                    nuevo_pais = st.text_input("Nuevo país",
                                               value=df_edit_src.loc[fila_edit, "Paises"],
                                               key="pais_edit_tab4")

                    nuevo_iemp = st.number_input("Nuevo IEMP",
                                                 value=_to_float(df_edit_src.loc[fila_edit,
                                                 "Índice de penetración en el mercado de exportación (IEMP)"]),
                                                 min_value=0.0)

                    nuevo_ioef = st.number_input("Nuevo IOEF",
                                                 value=_to_float(df_edit_src.loc[fila_edit,
                                                 "Índice de Libertad Económica (IOEF)"]),
                                                 min_value=0.0)

                    if st.button("Actualizar país (COMM)", key="btn_edit_tab4"):
                        ovr = st.session_state["comm_overrides"].copy()
                        ovr.at[fila_edit, "Paises"] = nuevo_pais
                        ovr.at[fila_edit, "Índice de penetración en el mercado de exportación (IEMP)"] = nuevo_iemp
                        ovr.at[fila_edit, "Índice de Libertad Económica (IOEF)"] = nuevo_ioef
                        st.session_state["comm_overrides"] = ovr
                        st.success("Registro actualizado en memoria.")
                        st.rerun()
                else:
                    st.info("No hay datos para editar aún.")

        # ===================== ELIMINAR =====================
        with col3:
            with st.expander("Eliminar país"):
                df_del_src = st.session_state.get("comm_overrides", pd.DataFrame())

                if not df_del_src.empty:
                    opciones_del = {r["Paises"]: i for i, r in df_del_src.iterrows()}
                    pais_sel_del = st.selectbox("Selecciona país a eliminar", list(opciones_del.keys()),
                                                key="pais_delete_tab4")
                    fila_delete = opciones_del[pais_sel_del]

                    if st.button("Eliminar país (COMM)", key="btn_delete_tab4"):
                        ovr = st.session_state["comm_overrides"].copy()
                        pais_eliminado = ovr.loc[fila_delete, "Paises"]
                        ovr = ovr.drop(fila_delete).reset_index(drop=True)
                        st.session_state["comm_overrides"] = ovr
                        st.warning(f"🗑️ País '{pais_eliminado}' eliminado.")
                        st.rerun()
                else:
                    st.info("No hay países en memoria para eliminar aún.")

        # ===================== TABLA PRINCIPAL =====================
        if "archivo_excel_bytes" not in st.session_state or st.session_state["archivo_excel_bytes"] is None:
            st.warning("Carga primero el archivo Excel en la pestaña 'Productos'.")
        else:
            try:
                archivo_excel = io.BytesIO(st.session_state["archivo_excel_bytes"])
                xls = pd.ExcelFile(archivo_excel, engine="openpyxl")

                hoja_paises = "Paises"
                hoja_pen = "Índice de penetración mercado"
                hoja_lib = "Indice de Libertad Económica"

                df_paises = pd.read_excel(xls, sheet_name=hoja_paises)
                df_pen = pd.read_excel(xls, sheet_name=hoja_pen)
                df_lib = pd.read_excel(xls, sheet_name=hoja_lib)


                def detectar_columna_pais(df):
                    for c in df.columns:
                        if isinstance(c, str) and ("pais" in c.lower() or "nombre" in c.lower()):
                            return c
                    return df.columns[0]

                c_pais_paises = detectar_columna_pais(df_paises)

                lista_paises = df_paises[c_pais_paises].dropna().astype(str).tolist()

                np.random.seed(42)
                tasas_auto = np.random.uniform(2.0, 10.0, len(lista_paises))

                df_comm = pd.DataFrame({
                    "Paises": lista_paises,
                    "Aranceles aduaneros por país de origen (CTCO)": [round(t, 2) for t in tasas_auto]
                })

                # ===================== MAPEO IEMP =====================
                c_pais_pen = detectar_columna_pais(df_pen)
                c_val_pen = next((c for c in df_pen.columns if c != c_pais_pen), None)
                df_pen.drop_duplicates(subset=[c_pais_pen], keep="last", inplace=True)

                df_comm["Índice de penetración en el mercado de exportación (IEMP)"] = (
                    df_comm["Paises"].map(df_pen.set_index(c_pais_pen)[c_val_pen])
                )

                # ===================== MAPEO IOEF =====================
                c_pais_lib = detectar_columna_pais(df_lib)
                c_val_lib = next((c for c in df_lib.columns if c != c_pais_lib), None)
                df_lib.drop_duplicates(subset=[c_pais_lib], keep="last", inplace=True)

                df_comm["Índice de Libertad Económica (IOEF)"] = (
                    df_comm["Paises"].map(df_lib.set_index(c_pais_lib)[c_val_lib])
                )

                # ===================== MERGE DE OVERRIDES =====================
                df_override = st.session_state.get("comm_overrides", pd.DataFrame())

                if not df_override.empty:
                    for _, row in df_override.iterrows():
                        pais = str(row["Paises"]).strip().lower()
                        mask = df_comm["Paises"].astype(str).str.lower().str.strip() == pais

                        if mask.any():
                            df_comm.loc[mask, "Índice de penetración en el mercado de exportación (IEMP)"] = row[
                                "Índice de penetración en el mercado de exportación (IEMP)"
                            ]
                            df_comm.loc[mask, "Índice de Libertad Económica (IOEF)"] = row[
                                "Índice de Libertad Económica (IOEF)"
                            ]

                        else:
                            nuevo = {
                                "Paises": row["Paises"],
                                "Aranceles aduaneros por país de origen (CTCO)": round(np.random.uniform(2, 10), 2),
                                "Índice de penetración en el mercado de exportación (IEMP)": row[
                                    "Índice de penetración en el mercado de exportación (IEMP)"
                                ],
                                "Índice de Libertad Económica (IOEF)": row[
                                    "Índice de Libertad Económica (IOEF)"
                                ]
                            }
                            df_comm = pd.concat([df_comm, pd.DataFrame([nuevo])], ignore_index=True)


                # CONVERSIÓN FUERTE A NÚMEROS (para que el orden funcione)

                cols = [
                    "Aranceles aduaneros por país de origen (CTCO)",
                    "Índice de penetración en el mercado de exportación (IEMP)",
                    "Índice de Libertad Económica (IOEF)"
                ]

                for c in cols:
                    df_comm[c] = pd.to_numeric(df_comm[c], errors="coerce")

                # ORDEN AUTOMÁTICO FINAL (CTCO ↓, IEMP ↑, IOEF ↑)

                df_comm = df_comm.sort_values(
                    by=[
                        "Aranceles aduaneros por país de origen (CTCO)",
                        "Índice de penetración en el mercado de exportación (IEMP)",
                        "Índice de Libertad Económica (IOEF)"
                    ],
                    ascending=[True, False, False],
                    na_position="last"
                ).reset_index(drop=True)

                # MOSTRAR TABLA CONSOLIDADA

                st.subheader("Tabla Comercial Consolidada (COMM)")
                st.dataframe(df_comm, use_container_width=True)

                # NORMALIZACIÓN (TU CÓDIGO NO SE TOCÓ)

                st.markdown("---")
                st.subheader("Tabla Comercial Normalizada (COMM)")

                A3 = 10

                def safe_norm(col, invertido=False):
                    col_min, col_max = col.min(), col.max()
                    if pd.isna(col_min) or pd.isna(col_max) or col_max == col_min:
                        return [None] * len(col)
                    if invertido:
                        return [round(A3 * (col_max - x) / (col_max - col_min), 2) if pd.notna(x) else None for x in
                                col]
                    else:
                        return [round(A3 * (x - col_min) / (col_max - col_min), 2) if pd.notna(x) else None for x in
                                col]

                df_norm = df_comm.copy()

                df_norm["CTCO_norm"] = safe_norm(
                    df_norm["Aranceles aduaneros por país de origen (CTCO)"],
                    invertido=True
                )
                df_norm["IEMP_norm"] = safe_norm(
                    df_norm["Índice de penetración en el mercado de exportación (IEMP)"]
                )
                df_norm["IOEF_norm"] = safe_norm(
                    df_norm["Índice de Libertad Económica (IOEF)"]
                )

                df_norm["COMM_Normalizado"] = (
                        df_norm["CTCO_norm"].fillna(0) * 0.5 +
                        df_norm["IEMP_norm"].fillna(0) * 0.3 +
                        df_norm["IOEF_norm"].fillna(0) * 0.2
                ).round(2)

                df_norm = df_norm.sort_values(by="COMM_Normalizado", ascending=False).reset_index(drop=True)

                st.caption("Ponderaciones: CTCO = 50 % | IEMP = 30 % | IOEF = 20 %")

                columnas_finales = ["Paises", "CTCO_norm", "IEMP_norm", "IOEF_norm", "COMM_Normalizado"]
                df_mostrar = df_norm[columnas_finales].rename(columns={"COMM_Normalizado": "COMM_total"})
                st.dataframe(df_mostrar, use_container_width=True)

                st.session_state["df_comm"] = df_norm[["Paises", "COMM_Normalizado"]].copy()

            except Exception as e:
                st.error(f"Error al procesar la hoja COMM: {e}")

    # ------------------ PESTAÑA 5 ------------------
    with tab5:
        st.title("4. Economía (ECON)")

        import os, io, datetime, openpyxl


        # ===============================================================
        # FUNCIÓN DE CONVERSIÓN (MISMA QUE TAB2)
        # ===============================================================
        def convertir_numero_seguro(valor):
            if valor is None:
                return None

            valor = str(valor).strip()

            if valor.lower() in ["none", "nan", "", "n/a", "null", "no data"]:
                return None

            simbolos = ["$", "USD", "₡", "CRC", "%"]
            for s in simbolos:
                valor = valor.replace(s, "")

            valor = valor.replace(" ", "")

            if valor.count(",") == 1 and valor.count(".") > 1:
                partes = valor.split(",")
                izquierda = partes[0].replace(".", "")
                valor = izquierda + "." + partes[1]

            elif valor.count(".") == 1 and valor.count(",") > 1:
                valor = valor.replace(",", "")

            elif valor.count(",") == 1 and valor.count(".") == 0:
                valor = valor.replace(",", ".")

            try:
                return float(valor)
            except:
                return None


        # ===============================================================
        # RUTA EXCEL MODIFICADO
        # ===============================================================
        ruta_excel = "2025-09-25T04-20_export_con_todos_los_paises_modificado.xlsx"
        if not os.path.exists(ruta_excel):
            st.error("No se encontró el archivo Excel modificado para ECON.")
            st.stop()

        # ===============================================================
        # CRUD ECON (INRA, INAN, DGDP)
        # ===============================================================
        st.markdown("### 🔧 Gestión de Datos Económicos (CRUD)")
        col1, col2, col3 = st.columns(3)

        # ------------------------ AÑADIR ------------------------
        with col1:
            with st.expander("Añadir país"):

                nuevo_pais = st.text_input("País nuevo", key="econ_add_pais")
                new_inra = st.number_input("INRA (Tasa de interés)", min_value=0.0, step=0.01, key="econ_add_inra")
                new_inan = st.number_input("INAN (Inflación anual)", min_value=0.0, step=0.01, key="econ_add_inan")
                new_dgdp = st.number_input("DGDP (Deuda/PIB)", min_value=0.0, step=0.01, key="econ_add_dgdp")

                if st.button("Guardar país ECON", key="btn_add_econ"):
                    wb = openpyxl.load_workbook(ruta_excel)

                    wb["Tasa de interés (INRA)"].append([nuevo_pais, new_inra])
                    wb["Inflación anual(INAN)"].append([nuevo_pais, new_inan])
                    wb["Relación deuda PIB(DGDP)"].append([nuevo_pais, new_dgdp])

                    try:
                        wb.save(ruta_excel)
                    except PermissionError:
                        base, ext = os.path.splitext(ruta_excel)
                        ts = datetime.datetime.now().strftime("%H%M%S")
                        copia = f"{base}_copia_{ts}{ext}"
                        wb.save(copia)
                        st.warning(f"Archivo abierto. Guardado como copia: {copia}")

                    st.success(f"País '{nuevo_pais}' añadido correctamente.")
                    st.rerun()

        # ===============================================================
        # CARGAR HOJAS ECON DESDE EL EXCEL MODIFICADO
        # ===============================================================
        xls = pd.ExcelFile(ruta_excel, engine="openpyxl")

        # --- INRA ---
        df_inra_raw = pd.read_excel(xls, sheet_name="Tasa de interés (INRA)")
        df_inra = df_inra_raw.rename(columns={"País": "Paises", "Último": "INRA"})
        df_inra["Paises"] = df_inra["Paises"].astype(str).str.strip()
        df_inra["INRA"] = df_inra["INRA"].apply(convertir_numero_seguro)

        # --- INAN ---
        df_inan_raw = pd.read_excel(xls, sheet_name="Inflación anual(INAN)")
        df_inan = df_inan_raw.rename(columns={"País": "Paises", "Inflación(% anual)": "INAN"})
        df_inan["Paises"] = df_inan["Paises"].astype(str).str.strip()
        df_inan["INAN"] = df_inan["INAN"].apply(convertir_numero_seguro)

        # --- DGDP ---
        df_dgdp_raw = pd.read_excel(xls, sheet_name="Relación deuda PIB(DGDP)")
        df_dgdp_raw.columns = df_dgdp_raw.columns.str.strip()

        df_dgdp = df_dgdp_raw.rename(columns={
            "país": "Paises",
            "pais": "Paises",
            "Relacion deuda(PIB": "DGDP"
        })

        df_dgdp["Paises"] = df_dgdp["Paises"].astype(str).str.strip()
        df_dgdp["DGDP"] = df_dgdp["DGDP"].apply(convertir_numero_seguro)

        # ------------------------ EDITAR Y ELIMINAR IGUAL ------------------------
        # (no se modifican, siguen igual que tu código)
        # -------------------------------------------------------------------------

        # ===============================================================
        # UNIR TABLAS
        # ===============================================================
        df_econ = df_inra.merge(df_inan, on="Paises", how="outer").merge(df_dgdp, on="Paises", how="outer")

        # ===============================================================
        # ORDENAR ORIGINAL
        # ===============================================================
        df_econ["completos"] = df_econ[["INRA", "INAN", "DGDP"]].notna().all(axis=1)
        df_econ = df_econ.sort_values("completos", ascending=False).drop(columns=["completos"])

        st.subheader("Tabla Económica (ECON) — Datos originales")
        df_econ_display = df_econ.rename(columns={
            "INRA": "Tasa de interés (INRA)",
            "INAN": "Inflación anual (INAN)",
            "DGDP": "Relación deuda PIB (DGDP)"
        })
        st.dataframe(df_econ_display, use_container_width=True)
        st.markdown("---")

        # ===============================================================
        # NORMALIZACIÓN — CON MÍNIMOS POSITIVOS
        # ===============================================================
        df_norm = df_econ.copy()

        RANGO = slice(4, 39)  # filas S5–S39

        # 👉 NUEVO: usar solo valores positivos para el mínimo
        min_inra = df_norm["INRA"].iloc[RANGO].apply(lambda x: x if x is not None and x > 0 else None).dropna().min()
        min_inan = df_norm["INAN"].iloc[RANGO].apply(lambda x: x if x is not None and x > 0 else None).dropna().min()
        min_dgdp = df_norm["DGDP"].iloc[RANGO].apply(lambda x: x if x is not None and x > 0 else None).dropna().min()


        def norm(valor, minimo):
            if valor is None or minimo is None:
                return None
            try:
                valor = float(valor)
            except:
                return None
            if valor <= 0:
                return None
            return round((10 * minimo) / valor, 4)


        df_norm["INRA_norm"] = df_norm["INRA"].apply(lambda x: norm(x, min_inra))
        df_norm["INAN_norm"] = df_norm["INAN"].apply(lambda x: norm(x, min_inan))
        df_norm["DGDP_norm"] = df_norm["DGDP"].apply(lambda x: norm(x, min_dgdp))

        # ===============================================================
        # PONDERACIÓN FINAL
        # ===============================================================
        P_INRA = 0.30
        P_INAN = 0.30
        P_DGDP = 0.40

        df_norm["Puntaje_ECON_Normalizado"] = (
                df_norm["INRA_norm"].fillna(0) * P_INRA +
                df_norm["INAN_norm"].fillna(0) * P_INAN +
                df_norm["DGDP_norm"].fillna(0) * P_DGDP
        ).round(4)

        # ===============================================================
        # ORDEN FINAL
        # ===============================================================
        df_norm["completos"] = df_norm[["INRA_norm", "INAN_norm", "DGDP_norm"]].notna().all(axis=1)

        df_rank = df_norm.sort_values(
            by=["completos", "Puntaje_ECON_Normalizado"],
            ascending=[False, False]
        ).drop(columns=["completos"])

        st.subheader("Tabla de Normalización Económica (ECON)")
        st.dataframe(
            df_rank[["Paises", "INRA_norm", "INAN_norm", "DGDP_norm", "Puntaje_ECON_Normalizado"]],
            use_container_width=True
        )

        st.session_state["df_econ"] = df_rank[["Paises", "Puntaje_ECON_Normalizado"]]

    # ------------------ PESTAÑA 6 ------------------
    with tab6:
        st.title("5. Política (POLI)")

        import openpyxl, os, datetime

        ruta_excel = os.path.abspath("2025-09-25T04-20_export_con_todos_los_paises_actualizado.xlsx")


        # ============================================================
        # Cargar datos actualizados desde el Excel
        # ============================================================
        def cargar_datos_excel():
            if not os.path.exists(ruta_excel):
                return pd.DataFrame(columns=[
                    "Paises",
                    "Índice de Estados Frágiles (FSI)",
                    "Informe sobre el riesgo (INRI)",
                    "Índice de Democracia (DEIN)"
                ])

            xls = pd.ExcelFile(ruta_excel, engine="openpyxl")

            def limpiar(df, nombre_columna):
                col_pais = next((c for c in df.columns if "pais" in str(c).lower()), df.columns[0])
                col_val = next((c for c in df.columns if c != col_pais), None)
                df = df[[col_pais, col_val]].rename(columns={col_pais: "Paises", col_val: nombre_columna})
                df["Paises"] = df["Paises"].astype(str).str.strip()
                df[nombre_columna] = pd.to_numeric(df[nombre_columna], errors="coerce")
                return df

            hojas = {h.lower(): h for h in xls.sheet_names}
            hoja_fsi = next((v for k, v in hojas.items() if "fsi" in k or "frag" in k), None)
            hoja_inri = next((v for k, v in hojas.items() if "inri" in k or "riesgo" in k), None)
            hoja_dein = next((v for k, v in hojas.items() if "dein" in k or "democ" in k), None)

            if not (hoja_fsi and hoja_inri and hoja_dein):
                st.error("No se detectaron correctamente las hojas (FSI, INRI, DEIN).")
                return pd.DataFrame()

            df_fsi = limpiar(pd.read_excel(xls, sheet_name=hoja_fsi), "Índice de Estados Frágiles (FSI)")
            df_inri = limpiar(pd.read_excel(xls, sheet_name=hoja_inri), "Informe sobre el riesgo (INRI)")
            df_dein = limpiar(pd.read_excel(xls, sheet_name=hoja_dein), "Índice de Democracia (DEIN)")

            df_poli = df_fsi.merge(df_inri, on="Paises", how="outer").merge(df_dein, on="Paises", how="outer")
            return df_poli


        df_poli = cargar_datos_excel()

        st.session_state["poli_overrides"] = df_poli.copy()


        # ============================================================
        # CRUD — escribir / eliminar valores en Excel
        # ============================================================
        def upsert_tres_hojas_excel(ruta, pais, fsi, inri, dein):
            wb = openpyxl.load_workbook(ruta)
            hojas = {h.lower(): h for h in wb.sheetnames}

            hoja_fsi = next((v for k, v in hojas.items() if "fsi" in k or "frag" in k), None)
            hoja_inri = next((v for k, v in hojas.items() if "inri" in k or "riesgo" in k), None)
            hoja_dein = next((v for k, v in hojas.items() if "dein" in k or "democ" in k), None)

            def escribir(ws, pais, valor):
                col_pais, col_val = 1, 2
                pais_norm = str(pais).strip().lower()
                fila = None
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=col_pais).value
                    if v and str(v).strip().lower() == pais_norm:
                        fila = r
                        break
                if fila:
                    ws.cell(row=fila, column=col_val, value=valor)
                else:
                    ws.append([pais, valor])

            escribir(wb[hoja_fsi], pais, fsi)
            escribir(wb[hoja_inri], pais, inri)
            escribir(wb[hoja_dein], pais, dein)

            wb.save(ruta)
            wb.close()
            return True


        def eliminar_pais_excel(ruta, pais):
            wb = openpyxl.load_workbook(ruta)
            hojas = {h.lower(): h for h in wb.sheetnames}

            hoja_fsi = next((v for k, v in hojas.items() if "fsi" in k or "frag" in k), None)
            hoja_inri = next((v for k, v in hojas.items() if "inri" in k or "riesgo" in k), None)
            hoja_dein = next((v for k, v in hojas.items() if "dein" in k or "democ" in k), None)

            def borrar(ws, pais):
                col_pais = 1
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=col_pais).value
                    if v and str(v).strip().lower() == pais.lower():
                        ws.delete_rows(r)
                        break

            for hoja in [hoja_fsi, hoja_inri, hoja_dein]:
                borrar(wb[hoja], pais)

            wb.save(ruta)
            wb.close()
            return True


        # ============================================================
        # CRUD UI
        # ============================================================
        st.markdown("### Gestión de Datos (Tabla POLI)")
        col1, col2, col3 = st.columns(3)

        # --------------------- AÑADIR ---------------------
        with col1:
            with st.expander("Añadir país"):
                pais_add = st.text_input("País", key="pais_add_tab6")
                fsi_add = st.number_input("Índice de Estados Frágiles (FSI)", min_value=0.0, key="fsi_add_tab6")
                inri_add = st.number_input("Informe sobre el riesgo (INRI)", min_value=0.0, key="inri_add_tab6")
                dein_add = st.number_input("Índice de Democracia (DEIN)", min_value=0.0, key="dein_add_tab6")

                if st.button("Guardar país (POLI)", key="btn_add_tab6"):
                    upsert_tres_hojas_excel(ruta_excel, pais_add, fsi_add, inri_add, dein_add)
                    st.success(f"País '{pais_add}' añadido correctamente.")
                    st.rerun()

        # --------------------- EDITAR ---------------------
        with col2:
            with st.expander("Editar país"):
                if not df_poli.empty:
                    pais_sel = st.selectbox("Selecciona país a editar", df_poli["Paises"].unique(),
                                            key="edit_select_tab6")
                    fila = df_poli[df_poli["Paises"] == pais_sel].iloc[0]


                    def _f(v):
                        return float(v) if not pd.isna(v) else 0.0


                    nuevo_pais = st.text_input("Nuevo país", pais_sel, key="edit_pais_tab6")
                    nuevo_fsi = st.number_input("Nuevo FSI", value=_f(fila["Índice de Estados Frágiles (FSI)"]))
                    nuevo_inri = st.number_input("Nuevo INRI", value=_f(fila["Informe sobre el riesgo (INRI)"]))
                    nuevo_dein = st.number_input("Nuevo DEIN", value=_f(fila["Índice de Democracia (DEIN)"]))

                    if st.button("Actualizar país (POLI)", key="btn_edit_tab6"):
                        upsert_tres_hojas_excel(ruta_excel, nuevo_pais, nuevo_fsi, nuevo_inri, nuevo_dein)
                        st.success("País actualizado correctamente.")
                        st.rerun()

        # --------------------- ELIMINAR ---------------------
        with col3:
            with st.expander("🗑️ Eliminar país"):
                if not df_poli.empty:
                    pais_sel = st.selectbox("Selecciona país a eliminar", df_poli["Paises"].unique(),
                                            key="delete_tab6")
                    if st.button("Eliminar país (POLI)", key="btn_delete_tab6"):
                        eliminar_pais_excel(ruta_excel, pais_sel)
                        st.warning(f"País '{pais_sel}' eliminado correctamente.")
                        st.rerun()

        # ============================================================
        # TABLA ORIGINAL ORDENADA
        # ============================================================
        if df_poli.empty:
            st.warning("No se encontraron datos en las hojas POLI.")
        else:
            df_poli["_faltantes"] = df_poli[
                ["Índice de Estados Frágiles (FSI)", "Informe sobre el riesgo (INRI)", "Índice de Democracia (DEIN)"]
            ].isna().sum(axis=1)

            df_poli = df_poli.sort_values(
                by=["_faltantes"],
                ascending=True
            ).drop(columns=["_faltantes"]).reset_index(drop=True)

            st.subheader("Tabla Política (POLI) — Datos originales")
            st.dataframe(df_poli, use_container_width=True)

            # ============================================================
            # NORMALIZACIÓN EXACTA SEGÚN FÓRMULAS EXCEL
            # ============================================================
            A3 = 10
            FSI_min = 14.3
            INRI_min = 1
            DEIN_max = 9.81

            df_norm = df_poli.copy()

            df_norm["FSI_norm"] = (A3 * FSI_min) / df_norm["Índice de Estados Frágiles (FSI)"]
            df_norm["INRI_norm"] = (A3 * INRI_min) / df_norm["Informe sobre el riesgo (INRI)"]
            df_norm["DEIN_norm"] = (A3 * df_norm["Índice de Democracia (DEIN)"]) / DEIN_max

            df_norm["FSI_norm"] = df_norm["FSI_norm"].round(2)
            df_norm["INRI_norm"] = df_norm["INRI_norm"].round(2)
            df_norm["DEIN_norm"] = df_norm["DEIN_norm"].round(2)

            df_norm["Puntaje_POLI_Normalizado"] = (
                    df_norm["FSI_norm"].fillna(0) * 0.35 +
                    df_norm["INRI_norm"].fillna(0) * 0.35 +
                    df_norm["DEIN_norm"].fillna(0) * 0.30
            ).round(2)

            df_rank = df_norm.copy()

            df_rank["_faltantes"] = df_rank[
                ["FSI_norm", "INRI_norm", "DEIN_norm"]
            ].isna().sum(axis=1)

            df_rank = df_rank.sort_values(
                by=["_faltantes", "Puntaje_POLI_Normalizado"],
                ascending=[True, False]
            ).drop(columns=["_faltantes"]).reset_index(drop=True)

            st.markdown("---")
            st.subheader("Tabla Política Normalizada (POLI)")
            st.caption("Ponderaciones: FSI = 35% | INRI = 35% | DEIN = 30%")

            st.dataframe(
                df_rank[[
                    "Paises",
                    "FSI_norm",
                    "INRI_norm",
                    "DEIN_norm",
                    "Puntaje_POLI_Normalizado"
                ]],
                use_container_width=True
            )

            st.session_state["df_poli"] = df_rank[["Paises", "Puntaje_POLI_Normalizado"]].copy()

    # ------------------ PESTAÑA 7 ------------------
    with tab7:
        st.title("6. Cultura (CULT)")

        # BOTONES CRUD — Añadir / Editar / Eliminar / Descargar

        st.markdown("###Gestión de Datos (Tabla CULT)")
        col1, col2, col3 = st.columns(3)

        # AÑADIR REGISTRO
        with col1:
            with st.expander("Añadir país"):
                pais_add = st.text_input("País", key="pais_add_tab7")
                glin_add = st.number_input("Índice de globalización (GLIN)", min_value=0.0, key="glin_add_tab7")
                cpci_add = st.number_input("Índice de Percepción de la Corrupción (CPCI)", min_value=0.0,
                                           key="cpci_add_tab7")
                cudi_add = st.number_input("Diferencia cultural (CUDI)", min_value=0.0, key="cudi_add_tab7")

                if st.button("Guardar país (CULT)", key="btn_add_tab7"):
                    nuevo = pd.DataFrame([[pais_add, glin_add, cpci_add, cudi_add]], columns=[
                        "Paises",
                        "Índice de globalización (GLIN)",
                        "Índice de Percepción de la Corrupción (CPCI)",
                        "Diferencia cultural (CUDI)"
                    ])
                    if "df_cult" not in st.session_state:
                        st.session_state["df_cult"] = nuevo
                    else:
                        st.session_state["df_cult"] = pd.concat(
                            [st.session_state["df_cult"], nuevo], ignore_index=True
                        )
                    st.success(f"País '{pais_add}' añadido correctamente.")
                    st.rerun()

        # EDITAR REGISTRO
        with col2:
            with st.expander("Editar país"):
                df_edit = st.session_state.get("df_cult", pd.DataFrame())
                if not df_edit.empty:
                    fila_edit = st.selectbox("Selecciona país a editar", df_edit.index.tolist(), key="fila_edit_tab7")
                    nuevo_pais = st.text_input("Nuevo país", value=df_edit.loc[fila_edit, "Paises"],
                                               key="pais_edit_tab7")
                    nuevo_glin = st.number_input(
                        "Nuevo GLIN",
                        value=float(df_edit.loc[fila_edit, "Índice de globalización (GLIN)"] or 0.0),
                        min_value=0.0,
                        key="glin_edit_tab7"
                    )
                    nuevo_cpci = st.number_input(
                        "Nuevo CPCI",
                        value=float(df_edit.loc[fila_edit, "Índice de Percepción de la Corrupción (CPCI)"] or 0.0),
                        min_value=0.0,
                        key="cpci_edit_tab7"
                    )
                    nuevo_cudi = st.number_input(
                        "Nuevo CUDI",
                        value=float(df_edit.loc[fila_edit, "Diferencia cultural (CUDI)"] or 0.0),
                        min_value=0.0,
                        key="cudi_edit_tab7"
                    )

                    if st.button("Actualizar país (CULT)", key="btn_edit_tab7"):
                        st.session_state["df_cult"].at[fila_edit, "Paises"] = nuevo_pais
                        st.session_state["df_cult"].at[fila_edit, "Índice de globalización (GLIN)"] = nuevo_glin
                        st.session_state["df_cult"].at[
                            fila_edit, "Índice de Percepción de la Corrupción (CPCI)"
                        ] = nuevo_cpci
                        st.session_state["df_cult"].at[fila_edit, "Diferencia cultural (CUDI)"] = nuevo_cudi
                        st.success("Registro actualizado correctamente.")
                        st.rerun()
                else:
                    st.info("No hay datos cargados para editar.")

        # ELIMINAR REGISTRO
        with col3:
            with st.expander("Eliminar país"):
                df_delete = st.session_state.get("df_cult", pd.DataFrame())
                if not df_delete.empty:
                    fila_delete = st.selectbox(
                        "Selecciona país a eliminar",
                        df_delete.index.tolist(),
                        key="fila_delete_tab7"
                    )
                    if st.button("Eliminar país (CULT)", key="btn_delete_tab7"):
                        pais_eliminado = df_delete.loc[fila_delete, "Paises"]
                        st.session_state["df_cult"] = df_delete.drop(fila_delete).reset_index(drop=True)
                        st.warning(f"País '{pais_eliminado}' eliminado correctamente.")
                        st.rerun()
                else:
                    st.info("No hay países para eliminar.")

        # DESCARGAR EXCEL ACTUALIZADO
        df_actual = st.session_state.get("df_cult", pd.DataFrame())
        if not df_actual.empty:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_actual.to_excel(writer, index=False, sheet_name="CULT_Actualizado")
            st.download_button(
                label="Descargar Excel actualizado",
                data=output.getvalue(),
                file_name="CULT_actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_tab7"
            )

        # BLOQUE NUEVO: CARGAR AUTOMÁTICAMENTE COLUMNA CUDI

        if "archivo_excel_bytes" in st.session_state and st.session_state["archivo_excel_bytes"] is not None:
            try:
                archivo_excel = io.BytesIO(st.session_state["archivo_excel_bytes"])
                xls = pd.ExcelFile(archivo_excel, engine="openpyxl")

                hoja_cudi = next((h for h in xls.sheet_names if "Diferencia cultural" in h), None)
                if hoja_cudi:
                    df_cudi = pd.read_excel(xls, sheet_name=hoja_cudi)
                    df_cudi.columns = df_cudi.columns.str.strip()

                    if "Paises" in df_cudi.columns and "Diferencia cultural (CUDI)" in df_cudi.columns:
                        if "df_cult" in st.session_state:
                            df_cult = st.session_state["df_cult"].copy()
                            df_cult = df_cult.drop(columns=["Diferencia cultural (CUDI)"], errors="ignore")
                            df_cult = df_cult.merge(
                                df_cudi[["Paises", "Diferencia cultural (CUDI)"]],
                                on="Paises",
                                how="left"
                            )
                            st.session_state["df_cult"] = df_cult.copy()
                            st.success("Columna 'Diferencia cultural (CUDI)' actualizada desde Excel.")

            except Exception as e:
                st.error(f"Error al leer la hoja 'Diferencia cultural (CUDI)': {e}")

        if "archivo_excel_bytes" not in st.session_state or st.session_state["archivo_excel_bytes"] is None:
            st.warning("Carga primero el archivo Excel en la pestaña 'Productos'.")
        else:
            try:
                import numpy as np

                archivo_excel = io.BytesIO(st.session_state["archivo_excel_bytes"])
                xls = pd.ExcelFile(archivo_excel, engine="openpyxl")

                hojas = [h.strip() for h in xls.sheet_names]
                hoja_glin = next(
                    (h for h in hojas if "Índice de globalización" in h or "Indice de globalizacion" in h),
                    None
                )
                hoja_cpci = next(
                    (h for h in hojas if "Índice de Percepción de la Corr" in h or "Indice de Percepcion" in h),
                    None
                )
                hoja_cudi = next(
                    (h for h in hojas if "Diferencia cultural" in h or "Diferencia Cultural" in h),
                    None
                )

                def leer_hoja(nombre_hoja, nombre_columna):
                    df_raw = pd.read_excel(xls, sheet_name=nombre_hoja, header=None)
                    fila_encabezado = None
                    for i, fila in df_raw.iterrows():
                        cad = " ".join(map(str, fila.astype(str))).lower()
                        if "pais" in cad or "país" in cad or "country" in cad:
                            fila_encabezado = i
                            break
                    if fila_encabezado is not None:
                        df = pd.read_excel(xls, sheet_name=nombre_hoja, header=fila_encabezado)
                    else:
                        df = pd.read_excel(xls, sheet_name=nombre_hoja)

                    df.columns = df.columns.astype(str).str.strip()
                    col_pais = next(
                        (c for c in df.columns if any(x in c.lower() for x in ["pais", "país", "country", "nombre"])),
                        None
                    )
                    col_valor = next((c for c in df.columns if c != col_pais), None)

                    if not col_pais or not col_valor:
                        return pd.DataFrame(columns=["Paises", nombre_columna])

                    df = df[[col_pais, col_valor]].rename(columns={col_pais: "Paises", col_valor: nombre_columna})
                    df["Paises"] = df["Paises"].astype(str).str.strip()
                    df[nombre_columna] = pd.to_numeric(df[nombre_columna], errors="coerce")
                    df = df.dropna(subset=["Paises"])
                    return df


                df_glin = leer_hoja(hoja_glin, "Índice de globalización (GLIN)")
                df_cpci = leer_hoja(hoja_cpci, "Índice de Percepción de la Corrupción (CPCI)")
                df_cudi = leer_hoja(hoja_cudi, "Diferencia cultural (CUDI)")

                df_cult = df_glin.merge(df_cpci, on="Paises", how="outer").merge(df_cudi, on="Paises", how="outer")

                # ---------- ORDEN PRIMERA TABLA (COMPLETOS ARRIBA) ----------
                df_cult["_faltantes"] = df_cult[[
                    "Índice de globalización (GLIN)",
                    "Índice de Percepción de la Corrupción (CPCI)",
                    "Diferencia cultural (CUDI)"
                ]].isna().sum(axis=1)

                df_cult = df_cult.sort_values(
                    by=["_faltantes"],
                    ascending=[True]
                ).drop(columns=["_faltantes"]).reset_index(drop=True)

                st.subheader("Tabla Cultural (CULT) — Datos originales combinados")
                st.dataframe(df_cult, use_container_width=True)

                # === NORMALIZACIÓN CON FÓRMULAS REALES DE EXCEL ===
                A3 = 10

                df_norm = df_cult.copy()

                # GLIN_norm = (A3 * GLIN) / MAX(GLIN_rango)
                serie_glin = pd.to_numeric(df_norm["Índice de globalización (GLIN)"], errors="coerce")
                max_glin = serie_glin.max(skipna=True)
                if pd.notna(max_glin) and max_glin > 0:
                    df_norm["GLIN_norm"] = serie_glin.apply(
                        lambda x: round(A3 * x / max_glin, 2) if pd.notna(x) and x > 0 else None
                    )
                else:
                    df_norm["GLIN_norm"] = None

                # CPCI_norm = (A3 * CPCI) / MAX(CPCI_rango)
                serie_cpci = pd.to_numeric(df_norm["Índice de Percepción de la Corrupción (CPCI)"], errors="coerce")
                max_cpci = serie_cpci.max(skipna=True)
                if pd.notna(max_cpci) and max_cpci > 0:
                    df_norm["CPCI_norm"] = serie_cpci.apply(
                        lambda x: round(A3 * x / max_cpci, 2) if pd.notna(x) and x > 0 else None
                    )
                else:
                    df_norm["CPCI_norm"] = None

                # CUDI_norm = (A3 * MIN(CUDI_rango)) / CUDI
                serie_cudi = pd.to_numeric(df_norm["Diferencia cultural (CUDI)"], errors="coerce")
                min_cudi = serie_cudi.min(skipna=True)
                if pd.notna(min_cudi) and min_cudi > 0:
                    df_norm["CUDI_norm"] = serie_cudi.apply(
                        lambda x: round(A3 * min_cudi / x, 2) if pd.notna(x) and x > 0 else None
                    )
                else:
                    df_norm["CUDI_norm"] = None

                # PONDERACIONES 30 / 50 / 20 (SUMAPRODUCTO)
                peso_glin, peso_cpci, peso_cudi = 0.30, 0.50, 0.20
                df_norm["Puntaje_CULT_Normalizado"] = (
                        df_norm["GLIN_norm"].fillna(0) * peso_glin +
                        df_norm["CPCI_norm"].fillna(0) * peso_cpci +
                        df_norm["CUDI_norm"].fillna(0) * peso_cudi
                ).round(2)

                # ---------- ORDEN SEGUNDA TABLA (COMPLETOS ARRIBA) ----------
                df_norm["_faltantes"] = df_norm[[
                    "GLIN_norm",
                    "CPCI_norm",
                    "CUDI_norm"
                ]].isna().sum(axis=1)

                df_rank = df_norm.sort_values(
                    by=["_faltantes", "Puntaje_CULT_Normalizado"],
                    ascending=[True, False]
                ).drop(columns=["_faltantes"]).reset_index(drop=True)

                st.markdown("---")
                st.subheader("Tabla Cultural Normalizada (CULT)")
                st.caption("Ponderaciones: GLIN = 30% | CPCI = 50% | CUDI = 20%")
                st.dataframe(
                    df_rank[[
                        "Paises", "GLIN_norm", "CPCI_norm", "CUDI_norm", "Puntaje_CULT_Normalizado"
                    ]],
                    use_container_width=True
                )

                st.session_state["df_cult"] = df_rank.copy()

            except Exception as e:
                st.error(f"Error al procesar la hoja CULT: {e}")

    # ------------------ PESTAÑA 8 ------------------
    with tab8:
        st.title("Visualización de Tablas Totales")

        tablas_requeridas = ["df_cost", "df_logi", "df_comm", "df_econ", "df_poli", "df_cult"]
        faltantes = [t for t in tablas_requeridas if t not in st.session_state]
        if faltantes:
            st.warning(
                f"Faltan las siguientes tablas normalizadas : {', '.join(faltantes)}. "
                "Primero abre cada pestaña (COST, LOGI, COMM, ECON, POLI, CULT) para que se carguen."
            )
        else:
            try:
                from functools import reduce
                import io

                # ------------------ Funciones auxiliares ------------------
                def normalizar_pais(df):
                    df["Paises"] = df["Paises"].astype(str).str.strip()
                    return df

                def detectar_columna(df, palabra):
                    for col in df.columns:
                        col_str = str(col).lower()
                        if palabra.lower() in col_str and "normalizado" in col_str:
                            return col
                    for col in df.columns:
                        if palabra.lower() in str(col).lower():
                            return col
                    return None

                # ------------------ Cargar DataFrames base ------------------
                df_cost = st.session_state["df_cost"]
                df_logi = st.session_state["df_logi"]
                df_comm = st.session_state["df_comm"]
                df_econ = st.session_state["df_econ"]
                df_poli = st.session_state["df_poli"]
                df_cult = st.session_state["df_cult"]

                # Detectar columnas normalizadas
                col_cost = detectar_columna(df_cost, "cost")
                col_logi = detectar_columna(df_logi, "logi")
                col_comm = detectar_columna(df_comm, "comm")
                col_econ = detectar_columna(df_econ, "econ")
                col_poli = detectar_columna(df_poli, "poli")
                col_cult = detectar_columna(df_cult, "cult")

                # Estandarizar nombres
                df_cost = df_cost[["Paises", col_cost]].rename(columns={col_cost: "1. Cost (COST)"})
                df_cost = normalizar_pais(df_cost)

                df_logi = df_logi[["Paises", col_logi]].rename(columns={col_logi: "2. Logistical (LOGI)"})
                df_logi = normalizar_pais(df_logi)

                df_comm = df_comm[["Paises", col_comm]].rename(columns={col_comm: "3. Commercial (COMM)"})
                df_comm = normalizar_pais(df_comm)

                df_econ = df_econ[["Paises", col_econ]].rename(columns={col_econ: "4. Economic (ECON)"})
                df_econ = normalizar_pais(df_econ)

                df_poli = df_poli[["Paises", col_poli]].rename(columns={col_poli: "5. Political (POLI)"})
                df_poli = normalizar_pais(df_poli)

                df_cult = df_cult[["Paises", col_cult]].rename(columns={col_cult: "6. Cultura (CULT)"})
                df_cult = normalizar_pais(df_cult)

                # ------------------ Unión de tablas ------------------
                df_total = reduce(
                    lambda left, right: pd.merge(left, right, on="Paises", how="outer"),
                    [df_cost, df_logi, df_comm, df_econ, df_poli, df_cult]
                )

                df_total["Paises"] = df_total["Paises"].astype(str).str.strip()

                columnas_numericas = [
                    "1. Cost (COST)", "2. Logistical (LOGI)", "3. Commercial (COMM)",
                    "4. Economic (ECON)", "5. Political (POLI)", "6. Cultura (CULT)"
                ]

                for col in columnas_numericas:
                    df_total[col] = pd.to_numeric(df_total[col], errors="coerce")

                total_paises = len(df_total)

                # ------------------ Países con datos completos ------------------
                df_completos = df_total.dropna(subset=columnas_numericas).copy()
                paises_con_datos = len(df_completos)

                # ------------------ Filtro de países productores (TAB1) ------------------
                if "productos_filtrados_tab1" in st.session_state:
                    df_prod = st.session_state["productos_filtrados_tab1"].copy()
                    df_prod["pais"] = df_prod["pais"].astype(str).str.strip()
                    paises_producto = df_prod["pais"].unique().tolist()
                    df_completos = df_completos[
                        df_completos["Paises"].astype(str).str.strip().isin(paises_producto)
                    ]

                if df_completos.empty:
                    st.warning(
                        "Ningún país coincide entre: datos completos + productor del producto (TAB1)."
                    )
                    st.stop()

                paises_incluidos = len(df_completos)
                paises_excluidos = total_paises - paises_incluidos

                # EXPANDERS DE PONDERACIÓN

                st.subheader("Ajuste manual de ponderaciones (IMSFE)")
                st.caption("Las ponderaciones solo se aplican cuando presiones el botón 'Aplicar ponderaciones'.")

                # Valores por defecto
                default_pesos_cat = {
                    "COST": 20, "LOGI": 20, "COMM": 15,
                    "ECON": 15, "POLI": 15, "CULT": 15
                }
                default_vars = {
                    "COST": {"PDPO": 35, "CDTI": 35, "CEBC": 30},
                    "LOGI": {"LPIN": 30, "CPT": 30, "ITTT": 40},
                    "COMM": {"CTCO": 50, "IEMP": 30, "IOEF": 20},
                    "ECON": {"INRA": 30, "INAN": 30, "DGDP": 40},
                    "POLI": {"FSI": 35, "INRI": 35, "DEIN": 30},
                    "CULT": {"GLIN": 30, "CPCI": 50, "CUDI": 20},
                }

                # Sesión para guardar pesos si el usuario los aplica
                if "ponderaciones_aplicadas" not in st.session_state:
                    st.session_state["ponderaciones_aplicadas"] = False
                    st.session_state["pesos_cat"] = default_pesos_cat.copy()
                    st.session_state["pesos_var"] = default_vars.copy()

                # ---------------------------
                # EXPANDERS DE CATEGORÍAS
                # ---------------------------
                with st.expander("Ponderación por categoría", expanded=False):
                    pesos_cat_mod = {}
                    suma = 0

                    cols = st.columns(3)
                    i = 0
                    for cat, val in st.session_state["pesos_cat"].items():
                        with cols[i % 3]:
                            nuevo = st.number_input(
                                f"{cat} (%)", value=float(val), min_value=0.0, max_value=100.0, step=1.0
                            )
                            pesos_cat_mod[cat] = nuevo
                        i += 1
                        suma += nuevo

                    if suma != 100:
                        st.error(f"La suma actual es {suma}%. Debe ser 100%.")
                    else:
                        st.success("La suma es 100%.")

                # ---------------------------
                # EXPANDERS DE VARIABLES INTERNAS
                # ---------------------------
                nuevos_vars = {}

                for categoria, variables in st.session_state["pesos_var"].items():
                    with st.expander(f"{categoria} — Variables internas", expanded=False):

                        suma_vars = 0
                        nuevos_vars[categoria] = {}

                        cols = st.columns(3)
                        idx = 0
                        for var, val in variables.items():
                            with cols[idx % 3]:
                                nuevo_val = st.number_input(
                                    f"{var} (%)", value=float(val), min_value=0.0, max_value=100.0, step=1.0
                                )
                                nuevos_vars[categoria][var] = nuevo_val
                                suma_vars += nuevo_val
                            idx += 1

                        if suma_vars != 100:
                            st.error(f"La suma de variables internas es {suma_vars}%. Debe ser 100%.")
                        else:
                            st.success("La suma es 100%.")

                # ---------------------------
                # BOTÓN APLICAR
                # ---------------------------
                if st.button("Aplicar ponderaciones"):
                    st.session_state["pesos_cat"] = pesos_cat_mod
                    st.session_state["pesos_var"] = nuevos_vars
                    st.session_state["ponderaciones_aplicadas"] = True
                    st.success("Ponderaciones aplicadas correctamente.")

                # Si no presionan el botón, usar valores por defecto
                pesos_cat = st.session_state["pesos_cat"]
                pesos_var = st.session_state["pesos_var"]

                # ---------------------------
                # CÁLCULO PUNTAJE TOTAL
                # ---------------------------
                df_completos["Puntaje Total"] = (
                        df_completos["1. Cost (COST)"] * (pesos_cat["COST"] / 100) +
                        df_completos["2. Logistical (LOGI)"] * (pesos_cat["LOGI"] / 100) +
                        df_completos["3. Commercial (COMM)"] * (pesos_cat["COMM"] / 100) +
                        df_completos["4. Economic (ECON)"] * (pesos_cat["ECON"] / 100) +
                        df_completos["5. Political (POLI)"] * (pesos_cat["POLI"] / 100) +
                        df_completos["6. Cultura (CULT)"] * (pesos_cat["CULT"] / 100)
                ).round(2)

                df_completos["Puntaje Total"] = df_completos["Puntaje Total"].clip(0, 10)

                st.session_state["df_total"] = df_completos.copy()

                # ---------------------------
                # TABLA GENERAL
                # ---------------------------
                st.subheader("Tabla General de Evaluación")
                st.caption(
                    f"Países incluidos: {paises_incluidos} / {total_paises} totales\n"
                    f"Países con datos completos antes de filtrar: {paises_con_datos}\n"
                    f"Países excluidos: {paises_excluidos}\n"
                    f"Pesos aplicados: {pesos_cat}"
                )

                st.dataframe(df_completos, use_container_width=True)

                # ---------------------------
                # TABLA DE RANKING
                # ---------------------------
                st.subheader("Ranking de Países")
                df_resumen = df_completos.sort_values(by="Puntaje Total", ascending=False).reset_index(drop=True)
                df_resumen.insert(0, "Ranking", range(1, len(df_resumen) + 1))
                st.dataframe(df_resumen[["Ranking", "Paises", "Puntaje Total"]], use_container_width=True)

            except Exception as e:
                st.error(f"Error al generar la tabla total: {e}")

    # ------------------ PESTAÑA 9 ------------------
    with tab9:
        st.title("Visualización de Gráficos Comparativos")

        if "df_total" not in st.session_state:
            st.warning("Primero genera la tabla total en la pestaña 'Visualización de Tablas Totales'.")
        else:
            import matplotlib.pyplot as plt
            import plotly.express as px
            import io

            # ------------------- CONFIGURACIÓN GENERAL -------------------
            df_total = st.session_state["df_total"].copy()

            # Verificar nombre correcto
            if "Puntaje Total" not in df_total.columns:
                st.error("ERROR: La columna correcta es 'Puntaje Total', pero no existe en df_total.")
                st.write("Columnas disponibles:", df_total.columns.tolist())
                st.stop()


            #                     GRÁFICO 1 — PLOTLY
            st.subheader("Comparativo de países mejor posicionados")

            df_top10 = df_total.sort_values(by="Puntaje Total", ascending=False).head(10)

            df_long = df_top10.melt(
                id_vars="Paises",
                value_vars=[
                    "1. Cost (COST)",
                    "2. Logistical (LOGI)",
                    "3. Commercial (COMM)",
                    "4. Economic (ECON)",
                    "5. Political (POLI)",
                    "6. Cultura (CULT)",
                    "Puntaje Total"
                ],
                var_name="Categoría",
                value_name="Puntaje"
            )

            fig = px.bar(
                df_long,
                x="Paises",
                y="Puntaje",
                color="Categoría",
                barmode="group",
                text="Puntaje",
                title="<b>Comparación de Puntajes</b>",
                color_discrete_sequence=px.colors.qualitative.Set2
            )

            fig.update_traces(
                texttemplate="%{text:.2f}",
                textposition="outside",
                cliponaxis=False,
                marker_line_width=0.8,
                marker_line_color="black"
            )

            fig.update_layout(
                height=700,
                font=dict(size=15, color="white"),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                xaxis_title="Países",
                yaxis_title="Puntaje Real"
            )

            st.plotly_chart(fig, use_container_width=True)

            #                     GRÁFICO 2 — MATPLOTLIB

            st.markdown("---")
            st.subheader("Puntaje Total")

            df_top30 = df_total.sort_values(by="Puntaje Total", ascending=False).head(30)

            fig2, ax2 = plt.subplots(figsize=(14, 6), dpi=180)
            fig2.patch.set_facecolor("#0d1117")
            ax2.set_facecolor("#0d1117")

            bars = ax2.bar(
                df_top30["Paises"],
                df_top30["Puntaje Total"],
                color="#00BFFF",
                edgecolor="white"
            )

            for bar in bars:
                h = bar.get_height()
                ax2.text(
                    bar.get_x() + bar.get_width() / 2,
                    h + 0.1,
                    f"{h:.2f}",
                    ha="center",
                    va="bottom",
                    rotation=90,
                    color="white",
                    fontsize=8,
                    fontweight="bold"
                )

            ax2.set_xticks(range(len(df_top30)))
            ax2.set_xticklabels(df_top30["Paises"], rotation=90, color="white", fontsize=8)

            ax2.set_ylabel("Puntaje Total", color="white", fontsize=12)
            ax2.set_title("Ranking — Puntaje Total (Top 30 países)", color="white", fontsize=14)
            ax2.grid(axis="y", linestyle="--", alpha=0.25, color="white")

            plt.tight_layout()
            st.pyplot(fig2)

            # DESCARGA GRÁFICO 2 (PNG)
            buffer_g2 = io.BytesIO()
            fig2.savefig(buffer_g2, format="png", dpi=300, bbox_inches="tight")
            st.download_button(
                "Descargar Gráfico 2 (PNG)",
                data=buffer_g2.getvalue(),
                file_name="Grafico_PuntajeTotal.png",
                mime="image/png"
            )

            #                     GRÁFICO 3 — IMSFE
            st.markdown("---")
            st.subheader("Comparativo IMSFE — Dimensiones y Puntaje Total (Modo Oscuro HD)")

            columnas_dim = [
                "1. Cost (COST)",
                "2. Logistical (LOGI)",
                "3. Commercial (COMM)",
                "4. Economic (ECON)",
                "5. Political (POLI)",
                "6. Cultura (CULT)"
            ]

            df_plot = df_total.dropna(subset=columnas_dim).copy()
            df_plot = df_plot.sort_values(by="Puntaje Total", ascending=False).head(30)

            fig3, ax3 = plt.subplots(figsize=(18, 7), dpi=200)
            fig3.patch.set_facecolor("#0d1117")
            ax3.set_facecolor("#0d1117")

            # Barras de Puntaje Total
            bars = ax3.bar(
                df_plot["Paises"],
                df_plot["Puntaje Total"],
                color="#00BFFF",
                edgecolor="white",
                width=0.6,
                label="Puntaje Total"
            )

            for bar in bars:
                h = bar.get_height()
                ax3.text(
                    bar.get_x() + bar.get_width() / 2,
                    h * 0.15,
                    f"{h:.2f}",
                    ha="center",
                    va="center",
                    rotation=90,
                    fontsize=12,
                    fontweight="bold",
                    color="white"
                )

            # Líneas
            colores = ["#FF5555", "#7CFF91", "#6A9EFF", "#00D4FF", "#E6A84F", "#CFCFCF"]
            marcadores = ["o", "^", "s", "D", "P", "X"]
            estilos = [":", "--", "-.", ":", "--", "-."]

            for col, color, marker, estilo in zip(columnas_dim, colores, marcadores, estilos):
                ax3.plot(
                    df_plot["Paises"],
                    df_plot[col],
                    color=color,
                    linestyle=estilo,
                    marker=marker,
                    markersize=6,
                    linewidth=1.2,
                    label=col
                )

            ax3.set_xticks(range(len(df_plot)))
            ax3.set_xticklabels(df_plot["Paises"], rotation=90, fontsize=11, color="white")

            ax3.set_ylim(0, 10)
            ax3.set_ylabel("Puntaje (0-10)", color="white", fontsize=12)
            ax3.set_title("Comparativo IMFSFE — Puntaje Total + Dimensiones", fontsize=15, color="white")
            ax3.tick_params(axis="y", colors="white", labelsize=10)
            ax3.legend(
                fontsize=10,
                loc="upper center",
                bbox_to_anchor=(0.5, 1.12),
                ncol=7,
                frameon=False,
                labelcolor="white"
            )

            ax3.grid(axis="y", linestyle="--", color="gray", alpha=0.3)

            st.pyplot(fig3)

            # DESCARGA GRÁFICO 3 (PNG)
            buffer_g3 = io.BytesIO()
            fig3.savefig(buffer_g3, format="png", dpi=300, bbox_inches="tight")
            st.download_button(
                "Descargar Gráfico 3 (PNG)",
                data=buffer_g3.getvalue(),
                file_name="Grafico_IMSFE.png",
                mime="image/png"
            )